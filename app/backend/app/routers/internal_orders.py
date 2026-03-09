from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.db import SessionLocal
from app import crud, models
from app.services import _enrich_orders_from_mp_list
from app.quote_templates import build_supplier_visible_payload
from app.integrations.lingxing_client import update_fbm_order, get_access_token, get_listing_search, get_shop_list, get_mws_order_detail
from app.config_store import get_lingxing_config
from datetime import datetime
import re
from typing import Any
import requests
import io
import os
from zoneinfo import ZoneInfo
from app.xlsx_utils import write_xlsx
from copy import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula.translate import Translator
import tempfile
import hashlib

router = APIRouter()
TEMPLATE_HEADERS = [
    "序列","出单日期","产品图","厘米","英寸","区域","工厂内部型号","包装/尾程方式","货款已付","店铺","订单编号","内部订单号","产品名","供应商","采购数量","单价","总价","1688运输单号","下单日期","供应商出货日期","到花街日期","开船日期","到港日期","单号","头程","尾程","套数","每箱套数","总箱数","货代箱规","计费重量","花街单价","头程运费总价","包材费","出库费","国内运费","其他","每套运费成本","毛重小于68kg","长cm","宽cm","高cm","长in＜80","宽in","高in","镑重量＜150lb","自算计费重","客户地址","联邦方式","联邦单号","联邦美金","反弹","反弹退回","购买配送","联邦人民币","总成本","回款","利润","售价","被退款","索赔额","SKU","发货日","送达日","卡派后台单号","oversize 130及165","周长＜419","出货图","POD","签收图","其他","快递/卡派账单月份"
]

SHOP_NAME_MAP = {
    "NAIROLET-US": "亚丰源",
    "煌明科技-US": "煌明",
    "简丽欧-US": "简丽欧",
    "晨阳铺货A-US": "晨阳",
    "TIZAZO-US": "口福轩",
    "Kadaligh-US": "维利安",
    "口服轩-CA": "口福轩",
    "爱瑞柔-US": "爱瑞柔",
    "路蔻尔-US": "路蔻尔",
    "聚乐-US": "聚乐",
    "译文-US": "译文",
}
DEFAULT_EXCHANGE_RATE = 7.0
_POLICY_LIMIT_MARKERS = (
    "应亚马逊政策要求，仅返回订购时间28天内数据",
)


def _map_shop_name(v: str | None) -> str:
    raw = str(v or "").strip()
    if not raw:
        return ""
    return SHOP_NAME_MAP.get(raw, raw)


def _get_exchange_rate(db: Session) -> float:
    cfg = crud.get_config(db, "internal_orders_settings")
    if cfg and isinstance(cfg.config_value, dict):
        try:
            rate = float(cfg.config_value.get("exchange_rate"))
            if rate > 0:
                return rate
        except Exception:
            pass
    return DEFAULT_EXCHANGE_RATE


def _needs_address_refresh(order: models.InternalOrder, ext_fields: dict) -> bool:
    if not order.platform_order_no:
        return False
    joined = "\n".join(str(v or "") for v in (ext_fields or {}).values())
    if any(marker in joined for marker in _POLICY_LIMIT_MARKERS):
        return True
    checks = [
        ext_fields.get("address_line1"),
        ext_fields.get("city"),
        ext_fields.get("state_or_region"),
        ext_fields.get("postal_code"),
        ext_fields.get("客户地址"),
        ext_fields.get("电话") or ext_fields.get("receiver_mobile") or ext_fields.get("receiver_tel"),
    ]
    if any(not _clean_text(v) for v in checks):
        return True
    # address_line2 是可选的，但若缺失则仍尝试回查，避免像 Unit A 这类二行地址长期漏掉
    return not _clean_text(ext_fields.get("address_line2"))


def _maybe_enrich_orders_for_view(db: Session, orders: list[models.InternalOrder]) -> None:
    if not orders:
        return
    pending = []
    for order in orders:
        ext_obj = crud.get_order_ext(db, order.id)
        ext_fields = dict(ext_obj.fields or {}) if ext_obj and isinstance(ext_obj.fields, dict) else {}
        if _needs_address_refresh(order, ext_fields):
            pending.append(order.platform_order_no)
    pending = [x for x in dict.fromkeys(pending) if x]
    if not pending:
        return
    cfg = get_lingxing_config(db)
    app_id = str(cfg.get("app_id") or "").strip()
    app_secret = str(cfg.get("app_secret") or "").strip()
    if not app_id or not app_secret:
        return
    token = get_access_token(app_id, app_secret)
    token_code = str(token.get("code"))
    if token_code not in ("0", "200"):
        return
    access_token = ((token.get("data") or {}).get("access_token") or "").strip()
    if not access_token:
        return
    for i in range(0, len(pending), 200):
        _enrich_orders_from_mp_list(db, access_token, app_id, "", pending[i:i+200])


def _to_ymd(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return f"{v.year}/{v.month}/{v.day}"
    s = str(v).strip()
    if not s:
        return ""
    dt = _parse_any_datetime(s)
    if dt is not None:
        return f"{dt.year}/{dt.month}/{dt.day}"
    return s


def _clean_text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if any(marker in s for marker in _POLICY_LIMIT_MARKERS):
        return ""
    return "" if s.lower() in ("none", "null", "nan") else s


def _country_text_zh(v: Any) -> str:
    code = _clean_text(v).upper()
    if not code:
        return ""
    mapping = {
        "US": "美国",
        "CA": "加拿大",
        "CN": "中国",
    }
    return mapping.get(code, code)


def _address_type_zh(v: Any) -> str:
    raw = _clean_text(v)
    if not raw:
        return ""
    if raw in ("1", "住宅", "Residential", "residential"):
        return "住宅"
    if raw in ("2", "办公", "商业地址", "商业", "Business", "business", "Office", "office"):
        return "办公"
    return raw


def _display_buyer_name(v: Any, receiver_name: Any = None) -> str:
    buyer = _clean_text(v)
    if not buyer:
        return ""
    parts = [p for p in re.split(r"\s+", buyer) if p]
    if not parts:
        return buyer
    receiver = _clean_text(receiver_name)
    # Amazon 页面常显示买家联系名的首词，如 "Daniel Waknine" -> "Daniel"
    # 原始全名仍保留在结构化字段 buyer_name 中，这里只做展示缩写。
    if len(parts) >= 2:
        return parts[0]
    return buyer


def _format_customer_address_block(fields: dict) -> str:
    name = _clean_text(fields.get("receiver_name") or fields.get("buyer_name") or fields.get("customer_name"))
    line1 = _clean_text(fields.get("address_line1"))
    line2 = _clean_text(fields.get("address_line2"))
    line3 = _clean_text(fields.get("address_line3"))
    district = _clean_text(fields.get("district"))
    doorplate = _clean_text(fields.get("doorplate_no"))
    city = _clean_text(fields.get("city") or fields.get("customer_city"))
    state = _clean_text(fields.get("state_or_region") or fields.get("customer_state"))
    postal = _clean_text(fields.get("postal_code") or fields.get("customer_zip"))
    country = _country_text_zh(fields.get("receiver_country_code") or fields.get("country_code"))
    addr_type = _address_type_zh(fields.get("address_type"))
    if not addr_type:
        addr_type = _address_type_zh(fields.get("address_type_name"))
    if not addr_type:
        addr_type = "住宅"
    buyer = _display_buyer_name(fields.get("buyer_name"), name)
    phone = _clean_text(fields.get("电话") or fields.get("receiver_mobile") or fields.get("receiver_tel"))

    city_line = f"{city}, {state}".strip(", ")
    if postal:
        city_line = f"{city_line} {postal}".strip()
    middle_lines = [x for x in [line2, line3] if x]
    tail_line = " ".join([x for x in [district, doorplate] if x]).strip()
    if tail_line:
        middle_lines.append(tail_line)

    rows = [
        name,
        line1,
        *middle_lines,
        city_line,
        country,
        f"地址类型:  {addr_type}" if addr_type else "",
        f"联系买家:\t{buyer}" if buyer else "",
        f"电话:\t{phone}" if phone else "",
    ]
    built = "\n".join([x for x in rows if x])
    if built:
        return built

    # fallback: clean existing raw address text
    raw = str(fields.get("客户地址") or fields.get("customer_address") or "").replace("\r", "\n").strip()
    if not raw:
        return ""
    raw = re.sub(r"(电话:\s*[+\d][+\d\-\s]*)\s*美国", r"\1\n美国", raw)
    raw = re.sub(r"(电话:\s*[+\d][+\d\-\s]*)\s*电话:\s*[+\d][+\d\-\s]*", r"\1", raw)
    raw = re.sub(r"(\d{5}(?:-\d{4})?)\s*([A-Z][A-Z][A-Z\s&'.-]{2,})", r"\1\n\2", raw)
    raw = re.sub(r"([A-Za-z][A-Za-z&'.\-\s]{2,})(\d{1,6}\s+[A-Za-z0-9#\-\s.]+)", r"\1\n\2", raw)
    parts = [x.strip() for x in raw.split("\n") if _clean_text(x)]
    seen_phone = False
    out = []
    for ln in parts:
        if ln.startswith("电话:"):
            if seen_phone:
                continue
            seen_phone = True
        out.append(ln)
    if country and country not in out:
        out.append(country)
    if addr_type and not any("地址类型:" in x for x in out):
        out.append(f"地址类型:  {addr_type}")
    if buyer and not any(x.startswith("联系买家:") for x in out):
        out.append(f"联系买家:\t{buyer}")
    if phone and not any(x.startswith("电话:") for x in out):
        out.append(f"电话:\t{phone}")
    return "\n".join(out)


def _to_zh_weekday(dt: datetime) -> str:
    return ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][dt.weekday()]


def _to_zh_tz(dt: datetime) -> str:
    return "PDT" if dt.astimezone(ZoneInfo("America/Los_Angeles")).dst() else "PST"


def _parse_any_datetime(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("T", " ").replace("Z", "").strip()
    s = re.sub(r"\s+", " ", s)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:19] if fmt.endswith("%S") else s[:10], fmt)
        except Exception:
            pass
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            return None
    return None


def _format_zh_date(v):
    dt = _parse_any_datetime(v)
    if not dt:
        return str(v or "").strip()
    la = dt.replace(tzinfo=ZoneInfo("America/Los_Angeles"))
    return f"{la.year}年{la.month}月{la.day}日{_to_zh_weekday(la)} {_to_zh_tz(la)}"


def _format_zh_date_range(v):
    s = str(v or "").strip()
    if not s:
        return ""
    if " 到 " in s:
        a, b = [x.strip() for x in s.split(" 到 ", 1)]
        return f"{_format_zh_date(a)} 到 {_format_zh_date(b)}"
    if " - " in s:
        a, b = [x.strip() for x in s.split(" - ", 1)]
        return f"{_format_zh_date(a)} 到 {_format_zh_date(b)}"
    dates = re.findall(r"\d{4}-\d{2}-\d{2}(?:[ T]\d{2}:\d{2}:\d{2})?", s)
    if len(dates) >= 2:
        return f"{_format_zh_date(dates[0])} 到 {_format_zh_date(dates[1])}"
    if len(dates) == 1:
        return _format_zh_date(dates[0])
    return s


def _split_product_name_and_code(v: str):
    s = str(v or "").strip()
    if not s:
        return "", ""
    lines = [x.strip() for x in s.splitlines() if x.strip()]
    if len(lines) >= 2:
        return lines[0], lines[1]
    code = _extract_product_code_segment(s)
    if code:
        return s.replace(code, "").strip(), code
    return s, ""


def _extract_cm_in_from_text(text: str):
    s = str(text or "").lower()
    # inch explicit
    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:in|inch|inches|英寸|''|\")", s)
    if m:
        try:
            inch = float(m.group(1))
            cm = inch * 2.54
            return str(int(round(cm))), f"{int(round(inch))}in"
        except Exception:
            pass
    # cm explicit
    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:cm|厘米)\b", s)
    if m:
        try:
            cm = float(m.group(1))
            inch = cm / 2.54
            return str(int(round(cm))), f"{int(round(inch))}in"
        except Exception:
            pass
    # encoded meter
    m = re.search(r"-(\d+(?:\.\d+)?)m\b", s)
    if m:
        try:
            cm = float(m.group(1)) * 100
            inch = cm / 2.54
            return str(int(round(cm))), f"{int(round(inch))}in"
        except Exception:
            pass
    # encoded cm
    m = re.search(r"-(\d+(?:\.\d+)?)cm\b", s)
    if m:
        try:
            cm = float(m.group(1))
            inch = cm / 2.54
            return str(int(round(cm))), f"{int(round(inch))}in"
        except Exception:
            pass
    return "", ""


def _normalized_product_full(raw_name: str, platform_order_no: str, ext_fields: dict):
    raw = str(raw_name or "").strip()
    if not raw:
        return ""
    if re.search(r"\n[A-Z0-9]{4,}-", raw, flags=re.I):
        return raw
    derived = _derive_cn_product_name(raw, platform_order_no, ext_fields or {})
    return derived or raw


def _extract_zip_from_address(addr: str):
    s = str(addr or "")
    m = re.search(r"\b(\d{5})(?:-\d{4})?\b", s)
    return m.group(1) if m else ""


def _derive_inches_text(cm_val, product_text: str):
    if cm_val not in (None, ""):
        try:
            cmf = float(str(cm_val).replace("cm", "").strip())
            if cmf > 0:
                return f"{int(round(cmf / 2.54))}in"
        except Exception:
            pass
    inc = _extract_inches_from_name(product_text or "")
    if inc:
        return f"{int(round(float(inc)))}in"
    return ""


def _enrich_groups_with_mws_detail(groups: list, detail_obj: dict):
    if not groups or not isinstance(detail_obj, dict):
        return groups
    item_list = detail_obj.get("item_list") or []
    if not isinstance(item_list, list) or not item_list:
        return groups
    detail_rows = []
    for d in item_list:
        if not isinstance(d, dict):
            continue
        sku = str(d.get("seller_sku") or d.get("sku") or d.get("local_sku") or "").strip()
        title = str(d.get("title") or d.get("product_name") or d.get("local_name") or "").strip()
        unit = d.get("unit_price_amount")
        if unit in (None, "", 0, "0", "0.0", "0.00"):
            item_price = d.get("item_price_amount")
            qty = d.get("quantity_ordered") or d.get("quantity") or 1
            try:
                unit = float(item_price) / max(float(qty), 1.0)
            except Exception:
                unit = None
        detail_rows.append({"sku": sku, "title": title, "unit": unit})

    used = set()
    for g in groups:
        sku = str(g.get("sku") or "").strip()
        # 1) exact sku match
        idx = -1
        if sku:
            for i, d in enumerate(detail_rows):
                if i in used:
                    continue
                if str(d.get("sku") or "").strip() and str(d.get("sku")).strip() == sku:
                    idx = i
                    break
        # 2) fallback first unused
        if idx < 0:
            for i, _ in enumerate(detail_rows):
                if i not in used:
                    idx = i
                    break
        if idx < 0:
            continue
        used.add(idx)
        pick = detail_rows[idx]
        if g.get("unit_price") in (None, "", 0, "0", "0.0", "0.00") and pick.get("unit") not in (None, ""):
            g["unit_price"] = pick.get("unit")
        if (not str(g.get("product_name") or "").strip()) and str(pick.get("title") or "").strip():
            g["product_name"] = pick.get("title")
    return groups


def _inches_number(v: str):
    s = str(v or "").strip().lower()
    m = re.search(r"(\d+(?:\.\d+)?)\s*in", s)
    if m:
        try:
            return float(m.group(1))
        except Exception:
            return None
    try:
        return float(s)
    except Exception:
        return None


def _cm_number(v: str):
    s = str(v or "").strip().lower().replace("cm", "")
    try:
        return float(s)
    except Exception:
        return None


def _download_image_to_file(url: str, cache_dir: str):
    u = str(url or "").strip()
    # Amazon thumbnail -> large image
    u = re.sub(r"_SL\d+_", "_SL800_", u, flags=re.I)
    if not u or not u.startswith(("http://", "https://")):
        return ""
    os.makedirs(cache_dir, exist_ok=True)
    key = hashlib.md5(u.encode("utf-8")).hexdigest()
    dst = os.path.join(cache_dir, f"{key}.img")
    if os.path.exists(dst) and os.path.getsize(dst) > 0:
        return dst
    try:
        resp = requests.get(u, timeout=20)
        if resp.status_code != 200 or not resp.content:
            return ""
        with open(dst, "wb") as f:
            f.write(resp.content)
        return dst
    except Exception:
        return ""


def _build_kapi_url(base_url: str, api_path: str) -> str:
    base = str(base_url or "").strip() or "https://tran.wedoexpress.com"
    # allow user to paste full endpoint URL in config
    if base.lower().endswith(api_path.lower()):
        return base
    return base.rstrip("/") + api_path


def _truncate(v, n: int):
    s = str(v or "")
    return s[:n] if len(s) > n else s


def _sanitize_kapi_query_payload(payload: dict) -> dict:
    if not isinstance(payload, dict):
        return {}
    p = dict(payload)
    shipper = dict(p.get("shipper") or {})
    receiver = dict(p.get("receiver") or {})
    # Provider side often limits reference_id lengths; keep these conservative.
    shipper["reference"] = _truncate(shipper.get("reference"), 64)
    shipper["customerOrderNo"] = _truncate(shipper.get("customerOrderNo"), 64)
    shipper["remark"] = _truncate(shipper.get("remark"), 50)
    receiver["remark"] = _truncate(receiver.get("remark"), 50)
    p["shipper"] = shipper
    p["receiver"] = receiver
    return p


def _normalize_order_status(v: str | None) -> str:
    s = str(v or "").strip().lower()
    if not s:
        return ""
    mapping = {
        "pending": "待审核",
        "pendingavailability": "未发货",
        "invoiceunconfirmed": "未发货",
        "unshipped": "未发货",
        "partiallyshipped": "部分发货",
        "shipped": "已发货",
        "canceled": "已取消/不发货",
        "cancelled": "已取消/不发货",
        "待审核": "待审核",
        "待发货": "未发货",
        "未发货": "未发货",
        "已发货": "已发货",
        "已取消": "已取消/不发货",
        "已取消/不发货": "已取消/不发货",
        "部分发货": "部分发货",
    }
    return mapping.get(s, str(v or "").strip())


def _extract_product_code_segment(v: str | None) -> str:
    txt = str(v or "")
    if not txt:
        return ""
    m = re.search(r"([A-Z0-9]{6,}-\d+(?:\.\d+)?m(?:\s*SLT)?)", txt, flags=re.I)
    if m:
        return m.group(1).strip()
    m2 = re.search(r"([A-Z0-9]{6,}-[A-Z0-9.]+(?:[-_\s]*SLT)?)", txt, flags=re.I)
    if m2:
        return m2.group(1).strip()
    return ""


def _extract_inches_from_name(name: str):
    text = (name or "").lower()
    vals = []
    for m in re.finditer(r"(\d+(?:\.\d+)?)\s*(?:in|inch|inches|英寸|''|\")", text):
        try:
            v = float(m.group(1))
            if v > 0:
                vals.append(v)
        except Exception:
            pass
    # cm -> in
    if not vals:
        for m in re.finditer(r"(\d+(?:\.\d+)?)\s*(?:cm|厘米)", text):
            try:
                cmv = float(m.group(1))
                if cmv > 0:
                    vals.append(cmv / 2.54)
            except Exception:
                pass
    # code like -1.37m / -90cm
    if not vals:
        m1 = re.search(r"-(\d+(?:\.\d+)?)m\b", text)
        if m1:
            try:
                vals.append(float(m1.group(1)) * 100 / 2.54)
            except Exception:
                pass
    if not vals:
        m2 = re.search(r"-(\d+(?:\.\d+)?)cm\b", text)
        if m2:
            try:
                vals.append(float(m2.group(1)) / 2.54)
            except Exception:
                pass
    return max(vals) if vals else None


def _infer_color_zh(ext_fields: dict, raw_name: str):
    # 1) listing tag color / explicit color fields
    color_src = (
        ext_fields.get("global_tags_color")
        or ext_fields.get("color")
        or ext_fields.get("颜色")
        or ext_fields.get("seller_color")
        or ""
    )
    low_src = str(color_src).lower()
    low_name = (raw_name or "").lower()

    # 2) parse from global_tags payload if present
    if not low_src:
        gt = ext_fields.get("global_tags")
        if isinstance(gt, list):
            low_src = " ".join([str(x) for x in gt]).lower()
        elif isinstance(gt, str):
            low_src = gt.lower()

    merged = f"{low_src} {low_name}"
    mapping = [
        ("wood", "木色"), ("oak", "木色"), ("walnut", "胡桃木色"),
        ("black", "黑色"), ("grey", "灰色"), ("gray", "灰色"),
        ("white", "白色"), ("gold", "金色"), ("blue", "蓝色"),
        ("green", "绿色"), ("beige", "米色"),
    ]
    for kw, zh in mapping:
        if kw in merged:
            return zh
    return "木色"


def _infer_feature_zh(raw_name: str):
    n = (raw_name or "").lower()
    parts = []
    if any(k in n for k in ["double", "dual", "2 sink", "double sink", "双盆"]):
        parts.append("双盆")
    # 形状词必须明确命中，避免把非圆镜款误判为圆镜
    if any(k in n for k in ["round mirror", "circular mirror", "circle mirror", "round-shaped", "圆镜"]):
        if any(k in n for k in ["split", "分割"]):
            parts.append("圆镜分割")
        elif any(k in n for k in ["double", "dual", "双"]):
            parts.append("双圆镜")
        else:
            parts.append("圆镜")
    if any(k in n for k in ["semi", "half", "半圆"]):
        parts.append("大半圆镜")
    if any(k in n for k in ["freestanding", "floor", "落地"]):
        parts.append("落地")
    if any(k in n for k in ["floating", "wall", "wall-mounted", "吊柜"]):
        parts.append("吊柜")
    if any(k in n for k in ["cabinet", "vanity", "柜"]):
        parts.append("柜体")
    if any(k in n for k in ["corian", "可丽耐"]):
        parts.append("可丽耐盆")
    # default
    if not parts:
        parts = ["简约", "柜体"]
    # 去重并保序
    out = []
    for p in parts:
        if p not in out:
            out.append(p)
    return "".join(out)


def _zip_to_region(zip_code: str) -> str:
    z = re.sub(r"\D", "", str(zip_code or ""))
    if not z:
        return ""
    d = int(z[0])
    if 0 <= d <= 3:
        return f"美东{d}"
    if 4 <= d <= 7:
        return f"美中{d}"
    return f"美西{d}"


def _derive_cn_product_name(raw_name: str, platform_order_no: str, ext_fields: dict):
    text = (raw_name or "").strip()
    if not text:
        return ""
    color_zh = _infer_color_zh(ext_fields, text)
    color_code_map = {"木色": "MS", "黑色": "HS", "灰色": "HUI", "白色": "BS", "金色": "JS", "蓝色": "LS", "绿色": "LVS", "米色": "MIS", "胡桃木色": "HTMS"}
    color_code = color_code_map.get(color_zh, "MS")
    feature_zh = _infer_feature_zh(text)
    style_code = "SCJY" if "简约" in feature_zh else "MSFG"
    inch = _extract_inches_from_name(text)
    size_m = (inch * 2.54 / 100.0) if inch else None
    size_txt_zh = f"{size_m:.2f}米" if size_m else ""
    size_txt_code = f"{size_m:.2f}m" if size_m else ""
    line1 = f"{color_zh}{feature_zh}浴室柜{size_txt_zh}".strip()
    seed = re.sub(r"\D", "", str(platform_order_no or "0"))[-2:] or "28"
    line2 = f"{color_code}{style_code}KFBV{seed}-{size_txt_code}" if size_txt_code else f"{color_code}{style_code}KFBV{seed}"
    return f"{line1}\n{line2}"


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def resolve_sid_list(access_token: str, app_id: str, cfg_sid: str | None) -> str:
    if not cfg_sid or str(cfg_sid).upper() == "ALL":
        shops = get_shop_list(access_token, app_id)
        if shops.get("code") == 0:
            return ",".join([str(s.get("sid")) for s in shops.get("data", []) if s.get("sid") is not None])
        return ""
    return str(cfg_sid)


def _extract_kapi_order_no(fields: dict) -> str:
    raw = (
        fields.get("卡派后台单号")
        or fields.get("orderNo#")
        or fields.get("orderNo")
        or ""
    )
    s = str(raw or "").strip()
    # allow formats like WTxxxx#, #WTxxxx, WTxxxx
    s = s.strip("#").strip()
    return s


@router.get("/")
def list_internal_orders(limit: int = 50, offset: int = 0, db: Session = Depends(get_db)):
    orders = crud.list_internal_orders(db, limit=limit, offset=offset)
    _maybe_enrich_orders_for_view(db, orders)
    items = []
    cfg = get_lingxing_config(db)
    app_id = str(cfg.get("app_id") or cfg.get("APP_ID") or "").strip()
    app_secret = str(cfg.get("app_secret") or cfg.get("APP_SECRET") or "").strip()
    _mws_token = None
    for o in orders:
        ext_obj = crud.get_order_ext(db, o.id)
        ext_fields = ext_obj.fields if ext_obj else {}
        def _bad_addr(v):
            if v is None:
                return True
            s = str(v).strip().lower()
            if not s:
                return True
            # treat placeholders like "None, None" / "null null" as empty
            normalized = s.replace(",", " ").replace("\n", " ")
            parts = [p for p in normalized.split() if p]
            return bool(parts) and all(p in ("none", "null", "nan") for p in parts)
        order_items = crud.get_order_items(db, o.id)
        if order_items:
            if not ext_fields.get("sku") and order_items[0].sku:
                ext_fields["sku"] = order_items[0].sku
            if not ext_fields.get("product_name") and order_items[0].product_name:
                ext_fields["product_name"] = order_items[0].product_name
        # 同SKU合并、不同SKU拆分（用于前端展示）
        grouped_lines = []
        if order_items:
            _g = {}
            for idx_i, it in enumerate(order_items, start=1):
                k = str(it.sku or "").strip() or f"__NO_SKU__{idx_i}"
                if k not in _g:
                    _g[k] = {
                        "sku": str(it.sku or "").strip(),
                        "quantity": int(it.quantity or 0),
                        "product_name": str(it.product_name or "").strip(),
                        "product_image": str(it.product_image or "").strip(),
                        "unit_price": it.unit_price,
                    }
                else:
                    _g[k]["quantity"] = int(_g[k].get("quantity") or 0) + int(it.quantity or 0)
            grouped_lines = list(_g.values())
            # 多SKU时优先用 mws/orderDetail 补齐每条SKU售价/标题
            try:
                need_enrich = len(grouped_lines) > 1 and any(
                    (ln.get("unit_price") in (None, "", 0, "0", "0.0", "0.00") or not str(ln.get("product_name") or "").strip())
                    for ln in grouped_lines
                )
                if need_enrich and app_id and app_secret and o.platform_order_no:
                    if not _mws_token:
                        tk = get_access_token(app_id, app_secret)
                        if tk.get("code") == 0:
                            _mws_token = tk.get("data", {}).get("access_token")
                    if _mws_token:
                        det = get_mws_order_detail(_mws_token, app_id, o.platform_order_no)
                        if det.get("code") == 0 and det.get("data"):
                            grouped_lines = _enrich_groups_with_mws_detail(grouped_lines, det.get("data")[0] or {})
            except Exception:
                pass
        if ext_fields.get("latest_ship_date") and not ext_fields.get("amz_ship"):
            ext_fields["amz_ship"] = ext_fields.get("latest_ship_date")
        if (ext_fields.get("earliest_delivery_date") or ext_fields.get("latest_delivery_date")) and not ext_fields.get("amz_deliver"):
            ext_fields["amz_deliver"] = f"{ext_fields.get('earliest_delivery_date','')} - {ext_fields.get('latest_delivery_date','')}".strip(" -")
        product_image = order_items[0].product_image if order_items else None
        if product_image == "/":
            product_image = None
        ext_img = ext_fields.get("product_image")
        if ext_img == "/":
            ext_img = None
        product_image = product_image or ext_img
        # map to template headers
        if o.purchase_time and not ext_fields.get("出单日期"):
            ext_fields["出单日期"] = _to_ymd(o.purchase_time)
        elif ext_fields.get("出单日期"):
            ext_fields["出单日期"] = _to_ymd(ext_fields.get("出单日期"))
        if product_image and not ext_fields.get("产品图"):
            ext_fields["产品图"] = product_image
        mapped_shop = _map_shop_name(o.shop_name)
        if mapped_shop and not ext_fields.get("店铺"):
            ext_fields["店铺"] = mapped_shop
        elif ext_fields.get("店铺"):
            ext_fields["店铺"] = _map_shop_name(ext_fields.get("店铺"))
        if o.platform_order_no and not ext_fields.get("订单编号"):
            ext_fields["订单编号"] = o.platform_order_no
        if o.internal_order_no and not ext_fields.get("内部订单号"):
            ext_fields["内部订单号"] = o.internal_order_no
        if (ext_fields.get("product_name") or ext_fields.get("产品名")) and not ext_fields.get("产品名"):
            ext_fields["产品名"] = ext_fields.get("product_name")
        if (ext_fields.get("purchase_qty") or ext_fields.get("采购数量")) and not ext_fields.get("采购数量"):
            ext_fields["采购数量"] = ext_fields.get("purchase_qty")
        # 价格口径：售价=平台销售价；单价=供应商报价
        if ext_fields.get("unit_price") and not ext_fields.get("售价"):
            ext_fields["售价"] = ext_fields.get("unit_price")
        if ext_fields.get("quoted_unit_price") and not ext_fields.get("单价"):
            ext_fields["单价"] = ext_fields.get("quoted_unit_price")
        # 单价由供应商报价回填，内部订单默认不自动填
        if ext_fields.get("单价") in (None, "", "0", 0, "0.0", "0.00"):
            ext_fields["单价"] = ""
        if (ext_fields.get("total_price") or ext_fields.get("总价")) and not ext_fields.get("总价"):
            ext_fields["总价"] = ext_fields.get("total_price")
        if o.tracking_no and not ext_fields.get("单号"):
            ext_fields["单号"] = o.tracking_no
        if ext_fields.get("customer_address") and not ext_fields.get("客户地址") and not _bad_addr(ext_fields.get("customer_address")):
            ext_fields["客户地址"] = ext_fields.get("customer_address")
        if _bad_addr(ext_fields.get("客户地址")):
            ext_fields["客户地址"] = ""
        if not ext_fields.get("客户地址"):
            name = ext_fields.get("receiver_name") or ext_fields.get("buyer_name") or ext_fields.get("customer_name")
            line1 = ext_fields.get("address_line1")
            line2 = ext_fields.get("address_line2")
            line3 = ext_fields.get("address_line3")
            district = ext_fields.get("district")
            doorplate = ext_fields.get("doorplate_no")
            city = ext_fields.get("city") or ext_fields.get("customer_city")
            state = ext_fields.get("state_or_region") or ext_fields.get("customer_state")
            zip5 = ext_fields.get("postal_code") or ext_fields.get("customer_zip")
            name = _clean_text(name)
            line1 = _clean_text(line1)
            line2 = _clean_text(line2)
            line3 = _clean_text(line3)
            district = _clean_text(district)
            doorplate = _clean_text(doorplate)
            city = _clean_text(city)
            state = _clean_text(state)
            zip5 = _clean_text(zip5)
            city_line = f"{city}, {state}".strip(", ")
            if zip5:
                city_line = f"{city_line} {zip5}".strip()
            mid = " ".join([x for x in [line2, line3, district, doorplate] if x])
            country_text = _country_text_zh(ext_fields.get("receiver_country_code") or ext_fields.get("country_code"))
            address_type_text = _address_type_zh(ext_fields.get("address_type"))
            buyer_contact = _clean_text(ext_fields.get("buyer_name"))
            phone_text = _clean_text(ext_fields.get("电话") or ext_fields.get("receiver_mobile") or ext_fields.get("receiver_tel"))
            addr = "\n".join(
                [x for x in [name, line1, mid, city_line, country_text] if x]
                + ([f"地址类型:  {address_type_text}"] if address_type_text else [])
                + ([f"联系买家: {buyer_contact}"] if buyer_contact else [])
                + ([f"电话: {phone_text}"] if phone_text else [])
            )
            if addr:
                ext_fields["客户地址"] = addr
        # 地址统一格式化：国家/地址类型/联系买家/电话，去重并规范换行
        formatted_addr = _format_customer_address_block(ext_fields)
        if formatted_addr:
            ext_fields["客户地址"] = formatted_addr
        if ext_fields.get("fedex_method") and not ext_fields.get("联邦方式"):
            ext_fields["联邦方式"] = ext_fields.get("fedex_method")
        if ext_fields.get("fedex_no") and not ext_fields.get("联邦单号"):
            ext_fields["联邦单号"] = ext_fields.get("fedex_no")
        if ext_fields.get("sku") and not ext_fields.get("SKU"):
            ext_fields["SKU"] = ext_fields.get("sku")
        if ext_fields.get("amz_ship") and not ext_fields.get("发货日"):
            ext_fields["发货日"] = ext_fields.get("amz_ship")
        if ext_fields.get("amz_deliver") and not ext_fields.get("送达日"):
            ext_fields["送达日"] = ext_fields.get("amz_deliver")
        if ext_fields.get("订单状态"):
            ext_fields["订单状态"] = _normalize_order_status(ext_fields.get("订单状态"))
        elif o.order_status:
            ext_fields["订单状态"] = _normalize_order_status(o.order_status)
        # 派生字段：英寸 / 区域 / 中文产品名编码
        raw_name = ext_fields.get("product_name") or ext_fields.get("产品名") or ""
        in_val = _extract_inches_from_name(raw_name)
        if in_val and not ext_fields.get("英寸"):
            ext_fields["英寸"] = f"{int(in_val) if float(in_val).is_integer() else round(in_val, 2)}IN"
        if not ext_fields.get("区域"):
            zip_code = ext_fields.get("postal_code") or ext_fields.get("customer_zip")
            region = _zip_to_region(zip_code)
            if not region:
                cc = str(ext_fields.get("receiver_country_code") or ext_fields.get("country_code") or "").upper()
                if cc and cc != "US":
                    region = _country_text_zh(cc)
            if region:
                ext_fields["区域"] = region
        persist_ext = {}
        if raw_name:
            derived_name = _derive_cn_product_name(raw_name, o.platform_order_no, ext_fields)
            if derived_name:
                ext_fields["产品名"] = derived_name
                if (ext_obj.fields if ext_obj else {}).get("产品名") != derived_name:
                    persist_ext["产品名"] = derived_name
            code_seg = _extract_product_code_segment(ext_fields.get("产品名") or derived_name)
            if code_seg:
                if not ext_fields.get("产品编码"):
                    ext_fields["产品编码"] = code_seg
                    if (ext_obj.fields if ext_obj else {}).get("产品编码") != code_seg:
                        persist_ext["产品编码"] = code_seg
                if not ext_fields.get("箱唛"):
                    ext_fields["箱唛"] = code_seg
                    if (ext_obj.fields if ext_obj else {}).get("箱唛") != code_seg:
                        persist_ext["箱唛"] = code_seg
        if persist_ext:
            crud.upsert_order_ext_bulk(db, o.id, persist_ext)
        if grouped_lines:
            for ln in grouped_lines:
                pname = ln.get("product_name") or ""
                cm_s, in_s = _extract_cm_in_from_text(pname)
                ln["厘米"] = cm_s or str(ext_fields.get("厘米") or ext_fields.get("长cm") or "")
                ln["英寸"] = in_s or str(ext_fields.get("英寸") or "")
                ln["产品名"] = _normalized_product_full(pname, o.platform_order_no, ext_fields)
        items.append({
            "id": o.id,
            "internal_order_no": o.internal_order_no,
            "platform_order_no": o.platform_order_no,
            "shop_name": mapped_shop,
            "order_status": _normalize_order_status(o.order_status),
            "tracking_no": o.tracking_no,
            "purchase_time": _to_ymd(o.purchase_time),
            "ext": ext_fields,
            "product_image": product_image,
            "line_groups": grouped_lines,
            "packages": [
                {
                    "length_cm": p.length_cm,
                    "width_cm": p.width_cm,
                    "height_cm": p.height_cm,
                    "length_in": p.length_in,
                    "width_in": p.width_in,
                    "height_in": p.height_in,
                    "weight_kg": p.weight_kg,
                    "weight_lb": p.weight_lb,
                    "billed_weight": p.billed_weight,
                }
                for p in crud.get_order_packages(db, o.id)
            ],
        })
    return {"items": items, "total": len(orders)}


@router.get("/settings")
def get_internal_order_settings(db: Session = Depends(get_db)):
    return {"exchange_rate": _get_exchange_rate(db)}


@router.post("/settings")
def set_internal_order_settings(payload: dict, db: Session = Depends(get_db)):
    try:
        rate = float(payload.get("exchange_rate"))
    except Exception:
        raise HTTPException(status_code=400, detail="invalid exchange_rate")
    if rate <= 0 or rate > 100:
        raise HTTPException(status_code=400, detail="exchange_rate out of range")
    crud.set_config(db, "internal_orders_settings", {"exchange_rate": rate})
    return {"ok": True, "exchange_rate": rate}


@router.post("/export-selected")
def export_selected_orders(payload: dict, db: Session = Depends(get_db)):
    ids = payload.get("order_ids") if isinstance(payload, dict) else None
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="missing order_ids")
    flat_orders = []
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", ".."))
    template_candidates = [
        "/Users/baicai/Downloads/26年1-6月亚马逊家具订单表测试.xlsx",
        "/Users/baicai/Downloads/internal_orders.xlsx",
        os.path.join(project_root, "docs", "inputs", "internal_orders.xlsx"),
        os.path.join(project_root, "app", "docs", "inputs", "internal_orders.xlsx"),
    ]
    template_path = next((p for p in template_candidates if os.path.exists(p)), template_candidates[1])
    cfg = get_lingxing_config(db)
    app_id = str(cfg.get("app_id") or cfg.get("APP_ID") or "").strip()
    app_secret = str(cfg.get("app_secret") or cfg.get("APP_SECRET") or "").strip()
    _mws_token = None
    exchange_rate = _get_exchange_rate(db)
    for oid in ids:
        try:
            order_id = int(oid)
        except Exception:
            continue
        o = crud.get_internal_order(db, order_id)
        if not o:
            continue
        ext_obj = crud.get_order_ext(db, o.id)
        ext = dict(ext_obj.fields or {}) if ext_obj and isinstance(ext_obj.fields, dict) else {}
        items = crud.get_order_items(db, o.id)
        grouped = {}
        if items:
            for idx_i, it in enumerate(items, start=1):
                sku_key = str(it.sku or "").strip() or f"__NO_SKU__{idx_i}"
                g = grouped.get(sku_key)
                if not g:
                    grouped[sku_key] = {
                        "sku": str(it.sku or "").strip(),
                        "qty": int(it.quantity or 0),
                        "product_name": str(it.product_name or "").strip(),
                        "unit_price": it.unit_price,
                        "image": str(it.product_image or "").strip(),
                    }
                else:
                    # 同 SKU 合并数量
                    g["qty"] = int(g.get("qty") or 0) + int(it.quantity or 0)
            try:
                need_enrich = len(grouped) > 1 and any(
                    (x.get("unit_price") in (None, "", 0, "0", "0.0", "0.00") or not str(x.get("product_name") or "").strip())
                    for x in grouped.values()
                )
                if need_enrich and app_id and app_secret and o.platform_order_no:
                    if not _mws_token:
                        tk = get_access_token(app_id, app_secret)
                        if tk.get("code") == 0:
                            _mws_token = tk.get("data", {}).get("access_token")
                    if _mws_token:
                        det = get_mws_order_detail(_mws_token, app_id, o.platform_order_no)
                        if det.get("code") == 0 and det.get("data"):
                            gl = list(grouped.values())
                            gl = _enrich_groups_with_mws_detail(gl, det.get("data")[0] or {})
                            grouped = {str(x.get("sku") or f'__NO_SKU__{i}'): x for i, x in enumerate(gl, start=1)}
            except Exception:
                pass
        else:
            grouped["__EXT__"] = {
                "sku": str(ext.get("SKU") or ext.get("sku") or "").strip(),
                "qty": int(ext.get("采购数量") or ext.get("purchase_qty") or 0),
                "product_name": str(ext.get("产品名") or ext.get("product_name") or "").strip(),
                "unit_price": ext.get("售价") or ext.get("unit_price") or "",
                "image": str(ext.get("产品图") or ext.get("product_image") or "").strip(),
            }
        customer_addr = _format_customer_address_block(ext) or str(ext.get("客户地址") or "").strip()
        region = str(ext.get("区域") or "").strip()
        if not region:
            zip5 = str(ext.get("postal_code") or "").strip() or _extract_zip_from_address(customer_addr)
            region = _zip_to_region(zip5)
        delivery = _format_zh_date_range(ext.get("送达日") or ext.get("amz_deliver") or "")
        ship = _format_zh_date(ext.get("发货日") or ext.get("latest_ship_date") or ext.get("amz_ship"))
        for g in grouped.values():
            ext_formatted_name = str(ext.get("产品名") or "").strip()
            multi_sku = len(grouped) > 1
            base_name = (g.get("product_name") or "").strip()
            if multi_sku:
                product_text = _normalized_product_full(base_name or ext.get("product_name") or ext_formatted_name, o.platform_order_no, ext)
            else:
                has_formatted_code = bool(re.search(r"\n[A-Z0-9]{4,}-", ext_formatted_name, flags=re.I))
                product_text = ext_formatted_name if has_formatted_code else _normalized_product_full(base_name or ext.get("product_name") or ext_formatted_name, o.platform_order_no, ext)
            pname_zh, pcode = _split_product_name_and_code(product_text)
            if not pcode:
                pcode = (
                    _extract_product_code_segment(ext_formatted_name)
                    or
                    _extract_product_code_segment(ext.get("产品名英文段"))
                    or _extract_product_code_segment(ext.get("箱唛"))
                    or _extract_product_code_segment(ext.get("Customer orderNo"))
                    or _extract_product_code_segment(ext.get("产品编码"))
                    or _extract_product_code_segment(ext.get("marks"))
                    or _extract_product_code_segment(ext.get("工厂内部型号"))
                    or _extract_product_code_segment(product_text)
                )
            cm = str(ext.get("厘米") or ext.get("长cm") or "").strip()
            cm_from_name, in_from_name = _extract_cm_in_from_text(product_text)
            if cm_from_name:
                cm = cm_from_name
            if not cm and pcode:
                m = re.search(r"-(\d+(?:\.\d+)?)m\b", str(pcode), flags=re.I)
                if m:
                    try:
                        cm = str(int(round(float(m.group(1)) * 100)))
                    except Exception:
                        cm = ""
            if not cm:
                cm = str(ext.get("厘米") or ext.get("长cm") or "").strip()
            inches_text = in_from_name or str(ext.get("英寸") or "").strip() or _derive_inches_text(cm, product_text)
            product_full = (f"{(pname_zh or product_text).strip()}\n{str(pcode or '').strip()}").strip()
            if product_full and not re.search(r"\n[A-Z0-9]{4,}-", product_full, flags=re.I):
                fallback_code = _extract_product_code_segment(product_full) or pcode
                if fallback_code:
                    product_full = f"{product_full.splitlines()[0]}\n{fallback_code}"
            qty = g.get("qty") or ext.get("采购数量") or ext.get("purchase_qty") or ""
            img = g.get("image") or ext.get("产品图") or ext.get("product_image") or ""
            order_row = {
                "序列": "",
                "出单日期": _to_ymd(ext.get("出单日期") or o.purchase_time),
                "产品图": img,
                "厘米": "",
                "英寸": inches_text,
                "区域": region,
                "工厂内部型号": ext.get("工厂内部型号") or ext.get("internal_factory_no") or "",
                "店铺": _map_shop_name(ext.get("店铺") or o.shop_name),
                "订单编号": ext.get("订单编号") or o.platform_order_no or "",
                "内部订单号": ext.get("内部订单号") or o.internal_order_no or "",
                "产品名_zh": pname_zh or product_text,
                "产品名_code": pcode or "",
                "产品名_full": product_full,
                "采购数量": qty,
                "单价": "",
                "售价": g.get("unit_price") or ext.get("售价") or ext.get("unit_price") or "",
                "SKU": g.get("sku") or ext.get("SKU") or ext.get("sku") or "",
                "单号": ext.get("单号") or o.tracking_no or "",
                "客户地址": customer_addr,
                "发货日": ship,
                "送达日": delivery,
                "订单状态": _normalize_order_status(ext.get("订单状态") or o.order_status or ""),
                "备注": ext.get("备注") or "水龙头单独打包",
            }
            # 4行块中这些列需要每行独立展示（不合并）
            for k in [
                "货代箱规", "计费重量", "花街单价", "头程运费总价", "毛重小于68kg",
                "长cm", "宽cm", "高cm", "长in\n＜80", "宽in", "高in", "镑重量\n＜150lb",
                "自算计费重", "oversize 130及165", "周长＜419",
            ]:
                if k not in order_row:
                    order_row[k] = ext.get(k, "")
            flat_orders.append(order_row)
    if not flat_orders:
        raise HTTPException(status_code=404, detail="no rows")
    if not os.path.exists(template_path):
        # fallback
        header = TEMPLATE_HEADERS
        rows = []
        for i, row_map in enumerate(flat_orders, start=1):
            row = []
            for h in header:
                if h == "序列":
                    row.append(str(i))
                elif h == "产品名":
                    row.append(row_map.get("产品名_full", ""))
                else:
                    row.append(str(row_map.get(h, "") or ""))
            rows.append(row)
        out_path = os.path.join("/tmp", f"internal_orders_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")
        write_xlsx(out_path, header, rows)
    else:
        wb = load_workbook(template_path)
        ws = wb.active
        # BU列以后不输出任何内容与线框
        max_keep_col = column_index_from_string("BT")
        if ws.max_column > max_keep_col:
            ws.delete_cols(max_keep_col + 1, ws.max_column - max_keep_col)
        # map header index
        header_map = {}
        for c in range(1, ws.max_column + 1):
            key = str(ws.cell(1, c).value or "").strip()
            if key:
                header_map[key] = c
        # detect style blocks in template (3-row and 4-row)
        template_start_3 = 2
        template_start_4 = None
        for mr in ws.merged_cells.ranges:
            if mr.min_col == 1 and mr.max_col == 1:
                span = mr.max_row - mr.min_row + 1
                if span == 3 and template_start_3 == 2:
                    template_start_3 = mr.min_row
                if span == 4 and template_start_4 is None:
                    template_start_4 = mr.min_row
        wb_tpl = load_workbook(template_path)
        ws_tpl = wb_tpl.active
        # clear old merges (except header)
        old_merges = list(ws.merged_cells.ranges)
        for mr in old_merges:
            if mr.min_row >= 2:
                ws.unmerge_cells(str(mr))
        # clear old data rows
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)
        # write all blocks with copied style
        def _extract_column_merge_spans(ws_src, start_row: int, block_size: int):
            spans = {c: 1 for c in range(1, ws_src.max_column + 1)}
            end_row = start_row + block_size - 1
            for m in ws_src.merged_cells.ranges:
                if m.min_row >= start_row and m.max_row <= end_row and m.min_col == m.max_col:
                    spans[m.min_col] = max(spans.get(m.min_col, 1), m.max_row - m.min_row + 1)
            return spans

        merge_spans_3 = _extract_column_merge_spans(ws_tpl, template_start_3, 3)
        merge_spans_4 = (
            _extract_column_merge_spans(ws_tpl, template_start_4, 4)
            if template_start_4 is not None
            else None
        )
        def _detect_formula_cols(start_row: int, block_size: int):
            cols = set()
            end_row = start_row + block_size - 1
            for rr in range(start_row, end_row + 1):
                for cc in range(1, ws_tpl.max_column + 1):
                    vv = ws_tpl.cell(rr, cc).value
                    if isinstance(vv, str) and vv.startswith("="):
                        cols.add(cc)
            return cols
        def _detect_formula_map(start_row: int, block_size: int):
            # key: (row_offset, col) -> template formula text
            fm = {}
            end_row = start_row + block_size - 1
            for rr in range(start_row, end_row + 1):
                off = rr - start_row
                for cc in range(1, ws_tpl.max_column + 1):
                    vv = ws_tpl.cell(rr, cc).value
                    if isinstance(vv, str) and vv.startswith("="):
                        fm[(off, cc)] = vv
            return fm
        formula_cols_3 = _detect_formula_cols(template_start_3, 3)
        formula_cols_4 = _detect_formula_cols(template_start_4, 4) if template_start_4 is not None else set()
        formula_map_3 = _detect_formula_map(template_start_3, 3)
        formula_map_4 = _detect_formula_map(template_start_4, 4) if template_start_4 is not None else {}
        # 业务规则：AB（每箱套数）不使用公式，按数据写入/留空
        ab_col = header_map.get("每箱套数")
        if ab_col:
            formula_cols_3.discard(ab_col)
            formula_cols_4.discard(ab_col)
            formula_map_3 = {(off, cc): f for (off, cc), f in formula_map_3.items() if cc != ab_col}
            formula_map_4 = {(off, cc): f for (off, cc), f in formula_map_4.items() if cc != ab_col}
        # 业务规则：BP/BQ/BR（出货图/POD/签收图）不输出模板图片公式，保持空白让用户后续上传
        suppress_formula_cols = {
            column_index_from_string("BP"),
            column_index_from_string("BQ"),
            column_index_from_string("BR"),
        }
        formula_cols_3 -= suppress_formula_cols
        formula_cols_4 -= suppress_formula_cols
        formula_map_3 = {(off, cc): f for (off, cc), f in formula_map_3.items() if cc not in suppress_formula_cols}
        formula_map_4 = {(off, cc): f for (off, cc), f in formula_map_4.items() if cc not in suppress_formula_cols}
        image_anchors = []
        def _merged_anchor(r: int, c: int):
            for mr in ws.merged_cells.ranges:
                if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                    return mr.min_row, mr.min_col
            return r, c

        def _set_value_safe(r: int, c: int, value: Any):
            if value in (None, ""):
                return
            ar, ac = _merged_anchor(r, c)
            cell = ws.cell(ar, ac)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = value
            elif cell.value in (None, ""):
                cell.value = value
            else:
                if str(value) not in str(cell.value):
                    cell.value = f"{cell.value}\n{value}"

        def _clear_cell_force(r: int, c: int):
            ar, ac = _merged_anchor(r, c)
            cell = ws.cell(ar, ac)
            cell.value = ""
            try:
                cell.hyperlink = None
            except Exception:
                pass

        current_row = 2
        for idx, data in enumerate(flat_orders, start=1):
            inc = _inches_number(data.get("英寸"))
            cmn = _cm_number(data.get("厘米"))
            use_4 = bool((inc is not None and inc <= 36) or (cmn is not None and cmn <= 91.44))
            block_size = 4 if use_4 else 3
            src_start = (template_start_4 if (use_4 and template_start_4 is not None) else template_start_3)
            start_row = current_row
            ws.insert_rows(start_row, amount=block_size)
            for r_off in range(block_size):
                # 当模板没有4行样式时，第4行复用第3行样式
                src_r = src_start + r_off
                if use_4 and template_start_4 is None and r_off == 3:
                    src_r = src_start + 2
                dst_r = start_row + r_off
                ws.row_dimensions[dst_r].height = ws_tpl.row_dimensions[src_r].height
                for c in range(1, ws.max_column + 1):
                    src = ws_tpl.cell(src_r, c)
                    dst = ws.cell(dst_r, c)
                    dst._style = copy(src._style)
                    dst.number_format = src.number_format
                    dst.font = copy(src.font)
                    dst.fill = copy(src.fill)
                    dst.border = copy(src.border)
                    dst.alignment = copy(src.alignment)
                    dst.protection = copy(src.protection)
                    # 带公式的模板必须保留并按目标行平移引用
                    v = src.value
                    if isinstance(v, str) and v.startswith("="):
                        try:
                            dst.value = Translator(v, origin=src.coordinate).translate_formula(dst.coordinate)
                        except Exception:
                            dst.value = v
                    else:
                        # 不复制模板静态值，避免把模板示例数据导出到新文件
                        dst.value = None
            # apply strict per-column merge rules from template
            if use_4 and merge_spans_4:
                col_spans = merge_spans_4
                formula_cols = formula_cols_4
            elif use_4 and not merge_spans_4:
                # fallback: extend 3-row rule to 4-row when no 4-row sample exists
                col_spans = {c: (4 if s == 3 else s) for c, s in merge_spans_3.items()}
                formula_cols = formula_cols_3
            else:
                col_spans = merge_spans_3
                formula_cols = formula_cols_3
            # 4行强制规则（用户指定）
            if use_4:
                for cc in range(column_index_from_string("T"), column_index_from_string("X") + 1):
                    col_spans[cc] = 3
                for cc in range(column_index_from_string("Z"), column_index_from_string("AC") + 1):
                    col_spans[cc] = 3
                for cc in range(column_index_from_string("AD"), column_index_from_string("AG") + 1):
                    col_spans[cc] = 1
                col_spans[column_index_from_string("AL")] = 3
                for cc in range(column_index_from_string("AM"), column_index_from_string("AU") + 1):
                    col_spans[cc] = 1
                for cc in range(column_index_from_string("AW"), column_index_from_string("AX") + 1):
                    col_spans[cc] = 3
                for cc in range(column_index_from_string("BN"), column_index_from_string("BO") + 1):
                    col_spans[cc] = 1
                col_spans[column_index_from_string("BR")] = 3
                col_spans[column_index_from_string("BT")] = 3
            for c in range(1, ws.max_column + 1):
                span = int(col_spans.get(c, 1) or 1)
                if span > 1:
                    ws.merge_cells(
                        start_row=start_row,
                        end_row=start_row + min(span, block_size) - 1,
                        start_column=c,
                        end_column=c,
                    )
            # common values (merged)
            def setv(name, value, row=start_row):
                c = header_map.get(name)
                if c:
                    # 公式列不直接写值，保留模板公式
                    if c in formula_cols:
                        return
                    _set_value_safe(row, c, value)

            setv("序列", idx)
            setv("出单日期", data.get("出单日期", ""))
            setv("产品图", data.get("产品图", ""))
            setv("厘米", "")
            setv("英寸", data.get("英寸", ""))
            setv("区域", data.get("区域", ""))
            setv("工厂内部型号", data.get("工厂内部型号", ""))
            setv("店铺", data.get("店铺", ""))
            setv("订单编号", data.get("订单编号", ""))
            setv("内部订单号", data.get("内部订单号", ""))
            setv("采购数量", data.get("采购数量", ""))
            setv("单价", data.get("单价", ""))
            setv("售价", data.get("售价", ""))
            setv("SKU", data.get("SKU", ""))
            setv("单号", data.get("单号", ""))
            setv("客户地址", data.get("客户地址", ""))
            setv("发货日", data.get("发货日", ""))
            setv("送达日", data.get("送达日", ""))
            if use_4:
                # 4行模板中不合并的列，按4行分别写值
                for cc in range(1, ws.max_column + 1):
                    if int(col_spans.get(cc, 1) or 1) != 1:
                        continue
                    if cc in formula_cols:
                        continue
                    hname = str(ws.cell(1, cc).value or "").strip()
                    v = data.get(hname, "")
                    if v in (None, ""):
                        continue
                    for rr in range(start_row, start_row + 4):
                        ws.cell(rr, cc).value = v
            # 强制回填公式（防止中间写值覆盖）
            fmap = formula_map_4 if use_4 else formula_map_3
            fsrc = template_start_4 if (use_4 and template_start_4 is not None) else template_start_3
            for (off, cc), f in fmap.items():
                if off >= block_size:
                    continue
                rr = start_row + off
                src_rr = fsrc + off
                try:
                    ws.cell(rr, cc).value = Translator(f, origin=f"{get_column_letter(cc)}{src_rr}").translate_formula(f"{get_column_letter(cc)}{rr}")
                except Exception:
                    ws.cell(rr, cc).value = f
            # 强制修正BD总成本公式：按块行数动态取最后一行的AL，避免4行漏算第4行
            c_bd = column_index_from_string("BD")
            if c_bd:
                r = start_row
                tail = start_row + block_size - 1
                ws.cell(r, c_bd).value = f"=Q{r}+AL{r}+BC{r}+AL{tail}"
            # 利润(BF)=回款(BE)*汇率-总成本(BD)，汇率支持配置
            c_bf = column_index_from_string("BF")
            c_be = column_index_from_string("BE")
            if c_bf and c_be and c_bd:
                rate_text = f"{exchange_rate:.4f}".rstrip("0").rstrip(".")
                if int(col_spans.get(c_bf, 1) or 1) == 1:
                    for rr in range(start_row, start_row + block_size):
                        ws.cell(rr, c_bf).value = f"={get_column_letter(c_be)}{rr}*{rate_text}-{get_column_letter(c_bd)}{rr}"
                else:
                    rr = start_row
                    ws.cell(rr, c_bf).value = f"={get_column_letter(c_be)}{rr}*{rate_text}-{get_column_letter(c_bd)}{rr}"
            # 4行补公式：模板部分共享公式在xml中会丢失，按列规则强制补齐R1-R4
            if use_4:
                c_ag = header_map.get("头程运费总价")
                c_ae = header_map.get("计费重量")
                c_af = header_map.get("花街单价")
                c_an = header_map.get("长cm")
                c_ao = header_map.get("宽cm")
                c_ap = header_map.get("高cm")
                c_am = header_map.get("毛重小于68kg")
                c_aq = header_map.get("长in\n＜80") or header_map.get("长in＜80")
                c_ar = header_map.get("宽in")
                c_as = header_map.get("高in")
                c_at = header_map.get("镑重量\n＜150lb") or header_map.get("镑重量＜150lb")
                c_au = header_map.get("自算计费重")
                c_bn = header_map.get("oversize 130及165")
                c_bo = header_map.get("周长＜419")
                for rr in range(start_row, start_row + 4):
                    if c_ag and c_ae and c_af:
                        ws.cell(rr, c_ag).value = f"={get_column_letter(c_ae)}{rr}*{get_column_letter(c_af)}{rr}"
                    if c_aq and c_an:
                        ws.cell(rr, c_aq).value = f"={get_column_letter(c_an)}{rr}/2.54"
                    if c_ar and c_ao:
                        ws.cell(rr, c_ar).value = f"={get_column_letter(c_ao)}{rr}/2.54"
                    if c_as and c_ap:
                        ws.cell(rr, c_as).value = f"={get_column_letter(c_ap)}{rr}/2.54"
                    if c_at and c_am:
                        ws.cell(rr, c_at).value = f"={get_column_letter(c_am)}{rr}*2.2046226"
                    if c_au and c_an and c_ao and c_ap:
                        ws.cell(rr, c_au).value = f"={get_column_letter(c_an)}{rr}*{get_column_letter(c_ao)}{rr}*{get_column_letter(c_ap)}{rr}/6000"
                    if c_bn and c_aq and c_ar and c_as:
                        ws.cell(rr, c_bn).value = f"={get_column_letter(c_aq)}{rr}+2*({get_column_letter(c_ar)}{rr}+{get_column_letter(c_as)}{rr})"
                    if c_bo and c_an and c_ao and c_ap:
                        ws.cell(rr, c_bo).value = f"={get_column_letter(c_an)}{rr}+2*({get_column_letter(c_ao)}{rr}+{get_column_letter(c_ap)}{rr})"
            # split lines
            name_col = header_map.get("产品名")
            marks_col = header_map.get("箱唛")
            note_col = header_map.get("备注")
            image_col = header_map.get("产品图")
            code = data.get("产品名_code", "")
            if name_col:
                _set_value_safe(start_row, name_col, data.get("产品名_full", ""))
            if marks_col:
                _set_value_safe(start_row, marks_col, code)
                _set_value_safe(start_row + 1, marks_col, code)
                _set_value_safe(start_row + 2, marks_col, f"{code} SLT".strip())
                if block_size == 4:
                    _set_value_safe(start_row + 3, marks_col, "")
            if note_col:
                _set_value_safe(start_row + 1, note_col, data.get("备注") or "水龙头单独打包")
            if image_col:
                img_url = data.get("产品图", "")
                # 不写URL文本，改为真正图片对象（Excel可点击查看/放大）
                _clear_cell_force(start_row, image_col)
                image_anchors.append((start_row, image_col, img_url))
                if header_map.get("箱唛"):
                    _clear_cell_force(start_row + 1, image_col)
                    _clear_cell_force(start_row + 2, image_col)
                    image_anchors.append((start_row + 1, image_col, img_url))
                    image_anchors.append((start_row + 2, image_col, img_url))
                    if block_size == 4:
                        _clear_cell_force(start_row + 3, image_col)
                        image_anchors.append((start_row + 3, image_col, img_url))
            # BP/BQ/BR列强制留空（不使用模板中的DISPIMG公式）
            for rr in range(start_row, start_row + block_size):
                for cc in suppress_formula_cols:
                    _clear_cell_force(rr, cc)
            current_row += block_size

        # remove trailing empty rows if any
        while ws.max_row >= current_row:
            ws.delete_rows(ws.max_row, 1)
        # embed image objects
        cache_dir = os.path.join(tempfile.gettempdir(), "ultimate_erp_img_cache")
        inserted = set()
        for r, c, u in image_anchors:
            if not u:
                continue
            ar, ac = _merged_anchor(r, c)
            pos = (ar, ac, u)
            if pos in inserted:
                continue
            inserted.add(pos)
            img_file = _download_image_to_file(u, cache_dir)
            if not img_file:
                continue
            try:
                ximg = XLImage(img_file)
                ximg.width = 800
                ximg.height = 800
                ximg.anchor = f"{get_column_letter(ac)}{ar}"
                ws.add_image(ximg)
            except Exception:
                continue
        out_path = os.path.join("/tmp", f"internal_orders_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(out_path)
        wb.close()
        wb_tpl.close()

    with open(out_path, "rb") as f:
        content = f.read()
    filename = f"internal_orders_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0",
    }
    return StreamingResponse(io.BytesIO(content), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get("/{internal_order_id}")
def get_internal_order(internal_order_id: int, db: Session = Depends(get_db)):
    order = crud.get_internal_order(db, internal_order_id)
    if not order:
        raise HTTPException(status_code=404, detail="order not found")
    _maybe_enrich_orders_for_view(db, [order])
    items = crud.get_order_items(db, internal_order_id)
    packages = crud.get_order_packages(db, internal_order_id)
    ext = crud.get_order_ext(db, internal_order_id)
    item_payload = [
        {
            "sku": i.sku,
            "product_name": i.product_name,
            "quantity": i.quantity,
            "product_image": i.product_image,
        }
        for i in items
    ]
    if not item_payload and ext and ext.fields:
        item_payload = [{
            "sku": ext.fields.get("sku"),
            "product_name": ext.fields.get("product_name"),
            "quantity": ext.fields.get("purchase_qty"),
            "product_image": ext.fields.get("product_image"),
        }]
    detail_ext = dict(ext.fields or {}) if ext and isinstance(ext.fields, dict) else {}
    detail_ext["订单状态"] = _normalize_order_status(detail_ext.get("订单状态") or order.order_status)

    return {
        "id": order.id,
        "internal_order_no": order.internal_order_no,
        "platform_order_no": order.platform_order_no,
        "shop_name": _map_shop_name(order.shop_name),
        "order_status": _normalize_order_status(order.order_status),
        "tracking_no": order.tracking_no,
        "purchase_time": _to_ymd(order.purchase_time),
        "ext": detail_ext,
        "items": item_payload,
        "packages": [
            {
                "length_cm": p.length_cm,
                "width_cm": p.width_cm,
                "height_cm": p.height_cm,
                "weight_kg": p.weight_kg,
            }
            for p in packages
        ],
    }


@router.delete("/{internal_order_id}")
def delete_internal_order(internal_order_id: int, db: Session = Depends(get_db)):
    order = crud.get_internal_order(db, internal_order_id)
    if not order:
        raise HTTPException(status_code=404, detail="order not found")
    # delete related rows first
    db.query(models.InternalOrderItem).filter(models.InternalOrderItem.internal_order_id == internal_order_id).delete()
    db.query(models.InternalOrderPackage).filter(models.InternalOrderPackage.internal_order_id == internal_order_id).delete()
    db.query(models.InternalOrderExt).filter(models.InternalOrderExt.internal_order_id == internal_order_id).delete()
    db.query(models.InternalOrder).filter(models.InternalOrder.id == internal_order_id).delete()
    db.commit()
    return {"ok": True, "id": internal_order_id}


@router.patch("/{internal_order_id}/fields")
def update_internal_order_fields(internal_order_id: int, payload: dict, db: Session = Depends(get_db)):
    field = payload.get("field")
    value = payload.get("value")
    if not field:
        # allow {key: value} payloads
        if len(payload.keys()) == 1:
            field = list(payload.keys())[0]
            value = payload.get(field)
        else:
            raise HTTPException(status_code=400, detail="missing field")
    crud.upsert_order_ext(db, internal_order_id, field, value)
    return {"ok": True}


@router.post("/")
def create_empty_internal_order(db: Session = Depends(get_db)):
    obj = crud.create_internal_order(db, {
        "internal_order_no": f"IO{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}",
        "order_status": "草稿",
        "purchase_time": datetime.utcnow(),
    })
    return {"id": obj.id}


def _parse_amazon_payload(payload: Any) -> dict:
    data = payload if isinstance(payload, dict) else {}
    # unwrap common nesting
    if "data" in data and isinstance(data["data"], dict):
        data = data["data"]
    item_list = data.get("item_list") or data.get("order_item") or []
    first_item = item_list[0] if item_list else {}
    return {
        "platform_order_no": data.get("amazon_order_id") or data.get("order_id") or data.get("merchant_order_id"),
        "order_status": data.get("order_status"),
        "purchase_time": data.get("purchase_date_local") or data.get("purchase_time") or data.get("purchase_date"),
        "tracking_no": data.get("tracking_number"),
        "shop_name": data.get("seller_name") or data.get("shop_name"),
        "item": {
            "sku": first_item.get("local_sku") or first_item.get("seller_sku") or first_item.get("sku"),
            "product_name": first_item.get("local_name") or first_item.get("product_name") or first_item.get("title"),
            "quantity": first_item.get("quantity_ordered") or first_item.get("quality") or first_item.get("quantity"),
            "product_image": first_item.get("small_image_url") or first_item.get("pic_url"),
            "asin": first_item.get("asin"),
        },
    }


@router.post("/import-amazon")
def import_amazon_order(payload: dict, db: Session = Depends(get_db)):
    parsed = _parse_amazon_payload(payload)
    if not parsed.get("platform_order_no"):
        raise HTTPException(status_code=400, detail="missing platform_order_no")
    order = crud.get_order_by_platform_no(db, parsed["platform_order_no"])
    if order:
        return {"id": order.id, "status": "exists"}
    purchase_time = parsed.get("purchase_time")
    if isinstance(purchase_time, str):
        try:
            purchase_time = datetime.fromisoformat(purchase_time.replace("Z", "+00:00"))
        except Exception:
            try:
                purchase_time = datetime.strptime(purchase_time, "%Y-%m-%d %H:%M:%S")
            except Exception:
                purchase_time = datetime.utcnow()
    if not purchase_time:
        purchase_time = datetime.utcnow()
    mapped = {
        "internal_order_no": f"IO{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}",
        "platform_order_no": parsed["platform_order_no"],
        "shop_name": parsed.get("shop_name"),
        "order_status": parsed.get("order_status") or "待审核",
        "purchase_time": purchase_time,
        "tracking_no": parsed.get("tracking_no"),
    }
    order = crud.create_internal_order(db, mapped)
    item = parsed.get("item") or {}
    if item.get("sku") or item.get("product_name"):
        crud.create_internal_order_item(db, order.id, {
            "sku": item.get("sku") or "",
            "product_name": item.get("product_name") or "",
            "quantity": item.get("quantity") or 1,
            "product_image": item.get("product_image") or "",
        })
    ext = {}
    if item.get("asin"):
        ext["asin"] = item.get("asin")
    if item.get("product_image"):
        ext["product_image"] = item.get("product_image")
    if item.get("product_name"):
        ext["product_name"] = item.get("product_name")
    if item.get("quantity"):
        ext["purchase_qty"] = item.get("quantity")
    if ext:
        crud.upsert_order_ext_bulk(db, order.id, ext)
        # auto fetch listing image by asin/sku
        cfg = get_lingxing_config(db)
        app_id = cfg.get("app_id")
        app_secret = cfg.get("app_secret")
        if app_id and app_secret:
            token = get_access_token(app_id, app_secret)
            if token.get("code") in (200, "200"):
                access_token = token.get("data", {}).get("access_token")
                sid = payload.get("sid") or payload.get("shop_id") or cfg.get("sid_list") or ""
                sid = resolve_sid_list(access_token, app_id, sid)
                search_field = None
                search_value = None
                if item.get("asin"):
                    search_field = "asin"
                    search_value = [item.get("asin")]
                elif item.get("sku"):
                    search_field = "seller_sku"
                    search_value = [item.get("sku")]
                if search_field and search_value:
                    res = get_listing_search(access_token, app_id, {
                        "sid": sid,
                        "is_pair": 1,
                        "is_delete": 0,
                        "search_field": search_field,
                        "search_value": search_value,
                        "exact_search": 1,
                        "store_type": 1,
                        "offset": 0,
                        "length": 15,
                    })
                    if res.get("code") == 0 and res.get("data"):
                        img = res["data"][0].get("small_image_url")
                        if img:
                            crud.upsert_order_ext(db, order.id, "product_image", img)
                            items = crud.get_order_items(db, order.id)
                            if items:
                                items[0].product_image = img
                                db.commit()
    return {"id": order.id, "status": "created"}


@router.post("/import-order-detail")
def import_order_detail(payload: dict, db: Session = Depends(get_db)):
    # Accept raw orderDetail response and update orders by amazon_order_id
    data = payload.get("data") if isinstance(payload, dict) else None
    if data is None and isinstance(payload, list):
        data = payload
    if not data:
        raise HTTPException(status_code=400, detail="missing data")
    updated = 0
    for d in data:
        amazon_order_id = d.get("amazon_order_id")
        if not amazon_order_id:
            continue
        order = crud.get_order_by_platform_no(db, amazon_order_id)
        if not order:
            continue
        item_list = d.get("item_list") or []
        if item_list:
            it = item_list[0]
            ext_update = {}
            if it.get("asin"):
                ext_update["asin"] = it.get("asin")
            sku_val = it.get("sku") or it.get("seller_sku")
            if sku_val:
                ext_update["sku"] = sku_val
            if it.get("product_name") or it.get("title"):
                ext_update["product_name"] = it.get("product_name") or it.get("title")
            img = it.get("pic_url")
            if img and img != "/":
                ext_update["product_image"] = img
            if d.get("latest_ship_date"):
                ext_update["latest_ship_date"] = d.get("latest_ship_date")
                ext_update["amz_ship"] = d.get("latest_ship_date")
            if d.get("earliest_delivery_date"):
                ext_update["earliest_delivery_date"] = d.get("earliest_delivery_date")
            if d.get("latest_delivery_date"):
                ext_update["latest_delivery_date"] = d.get("latest_delivery_date")
                ext_update["amz_deliver"] = f"{d.get('earliest_delivery_date','')} - {d.get('latest_delivery_date','')}".strip(" -")
            if ext_update:
                crud.upsert_order_ext_bulk(db, order.id, ext_update)
            items = crud.get_order_items(db, order.id)
            if items:
                if img and img != "/" and items[0].product_image in (None, "", "/"):
                    items[0].product_image = img
                    db.commit()
            else:
                crud.create_internal_order_item(db, order.id, {
                    "sku": it.get("sku") or it.get("seller_sku"),
                    "product_name": it.get("product_name") or it.get("title"),
                    "quantity": it.get("quantity_ordered"),
                    "unit_price": it.get("unit_price_amount"),
                    "currency": it.get("currency"),
                    "product_image": it.get("pic_url"),
                    "attachments": None,
                })
        updated += 1
    return {"updated": updated}


def _parse_cn_datetime(text: str):
    # e.g. 2026年1月27日周二 05:35 PST
    m = re.search(r"(\\d{4})年(\\d{1,2})月(\\d{1,2})日(?:\\S*)\\s*(\\d{1,2}:\\d{2})?", text or "")
    if not m:
        return None
    y, mo, d, tm = m.group(1), m.group(2), m.group(3), m.group(4)
    if not tm:
        tm = "00:00"
    try:
        return datetime(int(y), int(mo), int(d), int(tm.split(":")[0]), int(tm.split(":")[1]))
    except Exception:
        return None


ORDER_ID_RE = re.compile(r"订单编号：#\s*(\d{3}-\d{7}-\d{7})")
CITY_STATE_ZIP_RE = re.compile(r"^\s*([^,]+)\s*,\s*([A-Z]{2})\s*(\d{5})(?:-\d{4})?\s*$")


def _between(text: str, start: str, end: str) -> str:
    if not text:
        return ""
    sidx = text.find(start)
    if sidx < 0:
        return ""
    sidx += len(start)
    eidx = text.find(end, sidx)
    if eidx < 0:
        return ""
    return text[sidx:eidx]


def _between_any_end(text: str, start: str, ends: list) -> str:
    if not text:
        return ""
    sidx = text.find(start)
    if sidx < 0:
        return ""
    sidx += len(start)
    eidxs = []
    for e in ends:
        ei = text.find(e, sidx)
        if ei >= 0:
            eidxs.append(ei)
    if not eidxs:
        return ""
    return text[sidx:min(eidxs)]


def _split_orders_from_text(text: str):
    matches = list(ORDER_ID_RE.finditer(text))
    if not matches:
        return [text.strip()]
    pieces = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        pieces.append(text[start:end].strip())
    return [p for p in pieces if p]


def _normalize_space(text: str) -> str:
    return re.sub(r"[ \t]+", " ", (text or "").strip())


def _extract_address_lines(block: str):
    addr_block = _between_any_end(block, "配送地址", ["美国", "地址类型:", "联系买家:", "更多详情", "订单内容"])
    if not addr_block:
        return "", "", "", "", ""
    lines = [l.strip() for l in addr_block.splitlines() if l.strip()]
    if len(lines) < 2:
        return "", "", "", "", ""
    name = lines[0]
    city_state_zip = lines[-1]
    mid = lines[1:-1]
    line1 = mid[0] if len(mid) >= 1 else ""
    if not line1 and len(mid) >= 2:
        line1 = mid[1]
    city = state = zip5 = ""
    m = CITY_STATE_ZIP_RE.match(city_state_zip)
    if m:
        city, state, zip5 = m.group(1), m.group(2), m.group(3)
    return name, line1, city_state_zip, city, state, zip5


def _calc_region(zip5: str) -> str:
    if not zip5:
        return ""
    d = int(zip5[0])
    if 0 <= d <= 3:
        return f"美东{d}"
    if 4 <= d <= 7:
        return f"美中{d}"
    return f"美西{d}"


def _parse_packages(block: str):
    pkg_matches = list(re.finditer(r"包裹\s*(\d+)", block))
    if not pkg_matches:
        return []
    packages = []
    for i, m in enumerate(pkg_matches):
        start = m.start()
        end = pkg_matches[i + 1].start() if i + 1 < len(pkg_matches) else len(block)
        seg = block[start:end]
        carrier = re.search(r"承运人\s*([^\n]+)", seg)
        tracking = re.search(r"追踪编码\s*\n?\s*([0-9A-Z]+)", seg)
        service = re.search(r"配送服务\s*([^\n]+)", seg)
        ship_date = re.search(r"发货日期\s*([^\n]+)", seg)
        packages.append({
            "carrier": carrier.group(1).strip() if carrier else "",
            "tracking": tracking.group(1).strip() if tracking else "",
            "service": service.group(1).strip() if service else "",
            "ship_date": ship_date.group(1).strip() if ship_date else "",
        })
    return packages


def _extract_header_block(order_block: str) -> str:
    return _between(order_block, "订单一览", "订单内容") or ""


def _parse_top_dates(header_block: str):
    ship = _normalize_space(_between(header_block, "发货日期:", "送达日期:"))
    deliver = _normalize_space(_between(header_block, "送达日期:", "购买日期:"))
    purchase = _normalize_space(_between(header_block, "购买日期:", "配送服务:"))
    if not purchase:
        purchase = _normalize_space(re.search(r"购买日期[:：]?\s*([^\n]+)", header_block or "") .group(1)) if re.search(r"购买日期", header_block or "") else ""
    return ship, deliver, purchase


def _extract_phone(order_block: str) -> str:
    buyer_block = _between_any_end(order_block, "联系买家:", ["更多详情", "收税模型:"])
    m = re.search(r"电话:\s*([+\d\-\s]+)", buyer_block or "")
    if not m:
        m = re.search(r"电话:\s*([+\d\-\s]+)", order_block or "")
    if not m:
        return ""
    return re.sub(r"\D+", "", m.group(1) or "")


def _choose_package(packages):
    if not packages:
        return {}
    fedex = [p for p in packages if "fedex" in (p.get("carrier") or "").lower() and p.get("tracking")]
    if fedex:
        return fedex[-1]
    with_tracking = [p for p in packages if p.get("tracking")]
    if with_tracking:
        return with_tracking[-1]
    return packages[-1]


def _norm_carrier_name(v: str) -> str:
    s = (v or "").strip()
    up = s.upper()
    if "UPS" in up:
        return "UPS"
    if "FEDEX" in up:
        return "FedEx"
    return s


@router.post("/import-amazon-text")
def import_amazon_text(payload: dict, db: Session = Depends(get_db)):
    text = payload.get("text") or ""
    if not text.strip():
        raise HTTPException(status_code=400, detail="missing text")
    # normalize spaces/newlines
    text = text.replace("\u3000", " ").replace("\u00a0", " ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    blocks = _split_orders_from_text(text)
    created = 0
    updated = 0
    order_nos = []
    def extract_order_nos(src: str):
        return [m.group(1) for m in ORDER_ID_RE.finditer(src or "")]

    def find_order_no(src: str):
        m = ORDER_ID_RE.search(src or "")
        return m.group(1) if m else ""

    for b in blocks:
        order_no = find_order_no(b)
        if not order_no:
            continue
        order_nos.append(order_no)
        order = crud.get_order_by_platform_no(db, order_no)
        if not order:
            purchase = re.search(r"购买日期:\\s*([^\\n]+)", b)
            purchase_dt = _parse_cn_datetime(purchase.group(1)) if purchase else None
            mapped = {
                "internal_order_no": f"IO{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}",
                "platform_order_no": order_no,
                "order_status": "已发货",
                "purchase_time": purchase_dt or datetime.utcnow(),
            }
            order = crud.create_internal_order(db, mapped)
            created += 1
        else:
            updated += 1
        # extract fields
        header_block = _extract_header_block(b)
        ship_date_text, deliver_text, purchase_text = _parse_top_dates(header_block)
        phone_val = _extract_phone(b)
        asin = re.search(r"ASIN:\s*([A-Z0-9]+)", b)
        sku = re.search(r"SKU:\s*([^\n]+)", b)
        product = re.search(r"已发货\s*\n\s*([\s\S]+?)\nASIN:", b)
        if not product:
            product = re.search(r"商品名称\s*\n\s*([\s\S]+?)\nASIN:", b)
        qty = re.search(r"\n\s*(\d+)\s*US\$", b)
        status = re.search(r"订单内容[\s\S]+?\n([\u4e00-\u9fa5]{2,3})\s*\n图片", b)
        name, addr1, city_state_zip, city, state, zip5 = _extract_address_lines(b)
        customer_address = "\n".join([s for s in [name, addr1, city_state_zip] if s])
        packages = _parse_packages(b)
        chosen_pkg = _choose_package(packages)
        region = _calc_region(zip5)
        carrier = _norm_carrier_name(chosen_pkg.get("carrier", ""))
        service = (chosen_pkg.get("service") or "").strip()
        fedex_method = carrier or service
        if not carrier and chosen_pkg.get("tracking", "").upper().startswith("1Z"):
            carrier = "UPS"
            fedex_method = "UPS"
        ext_update = {
            "订单编号": order_no,
            "内部订单号": order.internal_order_no,
            "客户地址": customer_address,
            "区域": region,
            "出单日期": purchase_text,
            "发货日": ship_date_text,
            "送达日": deliver_text,
            "ASIN": asin.group(1).strip() if asin else "",
            "SKU": sku.group(1).strip() if sku else "",
            "产品名": _normalize_space(product.group(1)) if product else "",
            "联邦单号": chosen_pkg.get("tracking", ""),
            "联邦方式": fedex_method,
        }
        if phone_val:
            ext_update["电话"] = phone_val
            ext_update["receiver_mobile"] = phone_val
            if customer_address and f"电话: {phone_val}" not in customer_address:
                ext_update["客户地址"] = f"{customer_address}\n电话: {phone_val}"
        if qty:
            ext_update["采购数量"] = qty.group(1).strip()
        if status and status.group(1):
            order.order_status = status.group(1).strip()
        # update core columns for internal_orders table
        if purchase_text:
            order.purchase_time = _parse_cn_datetime(purchase_text) or order.purchase_time
        # map extra fields
        price = re.search(r"US\\$\\s*([\\d,]+\\.?\\d*)", b)
        if price:
            ext_update["售价"] = price.group(1).replace(",", "")
        if status and status.group(1):
            ext_update["订单状态"] = status.group(1).strip()
        # update all duplicated internal rows sharing the same platform_order_no
        same_orders = db.query(models.InternalOrder).filter(models.InternalOrder.platform_order_no == order_no).all()
        if not same_orders:
            same_orders = [order]
        for so in same_orders:
            crud.upsert_order_ext_bulk(db, so.id, ext_update)
        # item row
        items = crud.get_order_items(db, order.id)
        if not items:
            crud.create_internal_order_item(db, order.id, {
                "sku": ext_update.get("SKU"),
                "product_name": ext_update.get("产品名"),
                "quantity": 1,
                "product_image": "",
            })
    detected = order_nos[:]
    if created == 0 and updated == 0:
        # fallback: extract order numbers from whole text
        nums = extract_order_nos(text)
        for order_no in nums:
            order = crud.get_order_by_platform_no(db, order_no)
            if not order:
                mapped = {
                    "internal_order_no": f"IO{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}",
                    "platform_order_no": order_no,
                    "order_status": "已发货",
                    "purchase_time": datetime.utcnow(),
                }
                order = crud.create_internal_order(db, mapped)
                created += 1
            else:
                updated += 1
            detected.append(order_no)
    if created == 0 and updated == 0 and not detected:
        # still create a placeholder order to avoid losing pasted text
        mapped = {
            "internal_order_no": f"IO{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}",
            "platform_order_no": f"TEXT-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
            "order_status": "待审核",
            "purchase_time": datetime.utcnow(),
        }
        order = crud.create_internal_order(db, mapped)
        crud.upsert_order_ext_bulk(db, order.id, {"raw_text": text})
        created += 1
        detected.append(mapped["platform_order_no"])
    if not detected:
        detected = extract_order_nos(text)[:5]
    return {
        "created": created,
        "updated": updated,
        "orders": order_nos,
        "detected": detected,
        "blocks": len(blocks),
        "debug": {
            "text_head": text[:120],
            "match_count": len(extract_order_nos(text)),
            "digits_head": re.sub(r"\D", "", text)[:40],
            "parser_version": "20260210-3",
        }
    }

@router.post("/pull-order-detail")
def pull_order_detail(payload: dict, db: Session = Depends(get_db)):
    order_ids = payload.get("order_ids") or []
    if not order_ids:
        order_ids = [o.id for o in db.query(models.InternalOrder).all()]
    cfg = get_lingxing_config(db)
    app_id = cfg.get("app_id")
    app_secret = cfg.get("app_secret")
    if not app_id or not app_secret:
        raise HTTPException(status_code=400, detail="missing app_id/app_secret")
    token = get_access_token(app_id, app_secret)
    if token.get("code") not in (200, "200"):
        raise HTTPException(status_code=400, detail=str(token))
    access_token = token.get("data", {}).get("access_token")

    platform_ids = []
    for oid in order_ids:
        order = crud.get_internal_order(db, oid)
        if order and order.platform_order_no:
            platform_ids.append(order.platform_order_no)

    stats = {"total_orders": len(platform_ids), "detail_rows": 0, "asin": 0, "sku": 0, "img": 0, "item_empty": 0}

    for i in range(0, len(platform_ids), 200):
        batch = platform_ids[i:i+200]
        detail = get_mws_order_detail(access_token, app_id, batch)
        if detail.get("code") != 0:
            continue
        for d in detail.get("data", []) or []:
            stats["detail_rows"] += 1
            amazon_order_id = d.get("amazon_order_id")
            order = crud.get_order_by_platform_no(db, amazon_order_id)
            if not order:
                continue
            item_list = d.get("item_list") or []
            if not item_list:
                stats["item_empty"] += 1
                continue
            it = item_list[0]
            ext_update = {}
            if it.get("asin"):
                ext_update["asin"] = it.get("asin")
                stats["asin"] += 1
            sku_val = it.get("sku") or it.get("seller_sku")
            if sku_val:
                ext_update["sku"] = sku_val
                stats["sku"] += 1
            if it.get("product_name") or it.get("title"):
                ext_update["product_name"] = it.get("product_name") or it.get("title")
            img = it.get("pic_url")
            if img and img != "/":
                ext_update["product_image"] = img
                stats["img"] += 1
            if d.get("latest_ship_date"):
                ext_update["latest_ship_date"] = d.get("latest_ship_date")
                ext_update["amz_ship"] = d.get("latest_ship_date")
            if d.get("earliest_delivery_date"):
                ext_update["earliest_delivery_date"] = d.get("earliest_delivery_date")
            if d.get("latest_delivery_date"):
                ext_update["latest_delivery_date"] = d.get("latest_delivery_date")
                ext_update["amz_deliver"] = f"{d.get('earliest_delivery_date','')} - {d.get('latest_delivery_date','')}".strip(" -")
            if d.get("purchase_date_local"):
                ext_update["purchase_date_local"] = d.get("purchase_date_local")
            if ext_update:
                crud.upsert_order_ext_bulk(db, order.id, ext_update)
            items = crud.get_order_items(db, order.id)
            if not items:
                crud.create_internal_order_item(db, order.id, {
                    "sku": it.get("sku") or it.get("seller_sku"),
                    "product_name": it.get("product_name") or it.get("title"),
                    "quantity": it.get("quantity_ordered"),
                    "unit_price": it.get("unit_price_amount"),
                    "currency": it.get("currency"),
                    "product_image": it.get("pic_url"),
                    "attachments": None,
                })
            else:
                if items[0].sku in (None, "", " "):
                    items[0].sku = it.get("sku") or it.get("seller_sku")
                if items[0].product_name in (None, "", " "):
                    items[0].product_name = it.get("product_name") or it.get("title")
                if items[0].quantity in (None, 0):
                    items[0].quantity = it.get("quantity_ordered")
                if img and img != "/" and items[0].product_image in (None, "", "/"):
                    items[0].product_image = img
                db.commit()

    return {"ok": True, "stats": stats}


@router.post("/refresh-product-codes")
def refresh_product_codes(payload: dict, db: Session = Depends(get_db)):
    order_ids = payload.get("order_ids") or []
    if not order_ids:
        order_ids = [o.id for o in db.query(models.InternalOrder).all()]
    updated = 0
    skipped = 0
    for oid in order_ids:
        order = crud.get_internal_order(db, int(oid))
        if not order:
            continue
        ext = crud.get_order_ext(db, order.id)
        fields = dict(ext.fields or {}) if ext and isinstance(ext.fields, dict) else {}
        items = crud.get_order_items(db, order.id)
        raw_name = fields.get("product_name") or fields.get("产品名") or (items[0].product_name if items else "")
        if not raw_name:
            skipped += 1
            continue
        derived = _derive_cn_product_name(raw_name, order.platform_order_no, fields)
        code_seg = _extract_product_code_segment(derived) or _extract_product_code_segment(raw_name)
        patch = {}
        if derived and fields.get("产品名") != derived:
            patch["产品名"] = derived
        if code_seg and fields.get("产品编码") != code_seg:
            patch["产品编码"] = code_seg
        if code_seg and not fields.get("箱唛"):
            patch["箱唛"] = code_seg
        if patch:
            crud.upsert_order_ext_bulk(db, order.id, patch)
            updated += 1
        else:
            skipped += 1
    return {"ok": True, "total": len(order_ids), "updated": updated, "skipped": skipped}


@router.post("/refresh-pending-status")
def refresh_pending_status(payload: dict, db: Session = Depends(get_db)):
    order_ids = payload.get("order_ids") or []
    only_pending = bool(payload.get("only_pending", True))
    if not order_ids:
        base_q = db.query(models.InternalOrder)
        if only_pending:
            base_q = base_q.filter(
                models.InternalOrder.order_status.in_(
                    ["待审核", "Pending", "pending", "未发货", "待发货", "Unshipped", "unshipped"]
                )
            )
        order_ids = [x.id for x in base_q.all()]
    if not order_ids:
        return {"ok": True, "total": 0, "updated": 0}

    cfg = get_lingxing_config(db)
    app_id = cfg.get("app_id")
    app_secret = cfg.get("app_secret")
    if not app_id or not app_secret:
        raise HTTPException(status_code=400, detail="missing app_id/app_secret")
    token = get_access_token(app_id, app_secret)
    if token.get("code") not in (200, "200"):
        raise HTTPException(status_code=400, detail=str(token))
    access_token = token.get("data", {}).get("access_token")

    uniq_order_nos = []
    id_by_no = {}
    for oid in order_ids:
        o = crud.get_internal_order(db, int(oid))
        if not o or not o.platform_order_no:
            continue
        no = str(o.platform_order_no).strip()
        if not no:
            continue
        if no not in id_by_no:
            id_by_no[no] = []
            uniq_order_nos.append(no)
        id_by_no[no].append(o.id)

    updated = 0
    detail_hits = 0
    for i in range(0, len(uniq_order_nos), 200):
        batch = uniq_order_nos[i:i + 200]
        detail = get_mws_order_detail(access_token, app_id, batch)
        if detail.get("code") != 0:
            continue
        for d in detail.get("data", []) or []:
            order_no = str(d.get("amazon_order_id") or "").strip()
            if not order_no or order_no not in id_by_no:
                continue
            detail_hits += 1
            new_status = _normalize_order_status(d.get("order_status"))
            if not new_status:
                continue
            for oid in id_by_no[order_no]:
                o = crud.get_internal_order(db, oid)
                if not o:
                    continue
                changed = False
                if (o.order_status or "") != new_status:
                    o.order_status = new_status
                    o.updated_at = datetime.utcnow()
                    changed = True
                ext = crud.get_order_ext(db, oid)
                fields = dict(ext.fields or {}) if ext and isinstance(ext.fields, dict) else {}
                if fields.get("订单状态") != new_status:
                    fields["订单状态"] = new_status
                    crud.upsert_order_ext_bulk(db, oid, fields)
                    changed = True
                if changed:
                    db.commit()
                    updated += 1

    return {
        "ok": True,
        "total": len(order_ids),
        "order_nos": len(uniq_order_nos),
        "detail_hits": detail_hits,
        "updated": updated,
    }


@router.post("/deduplicate")
def deduplicate_orders(payload: dict, db: Session = Depends(get_db)):
    order_ids = payload.get("order_ids") or []
    target_ids = set(int(x) for x in order_ids) if order_ids else set()

    q = db.query(models.InternalOrder).filter(models.InternalOrder.platform_order_no.isnot(None))
    all_rows = [o for o in q.all() if str(o.platform_order_no or "").strip()]
    groups = {}
    for o in all_rows:
        key = str(o.platform_order_no).strip()
        groups.setdefault(key, []).append(o)

    removed_ids = []
    kept = 0
    groups_count = 0
    for _no, rows in groups.items():
        if len(rows) <= 1:
            continue
        if target_ids and not any(r.id in target_ids for r in rows):
            continue
        groups_count += 1

        def score(order_obj):
            ext = crud.get_order_ext(db, order_obj.id)
            fields = ext.fields if ext and isinstance(ext.fields, dict) else {}
            s = 0
            if fields.get("客户地址"): s += 3
            if fields.get("电话"): s += 2
            if fields.get("ASIN"): s += 2
            if fields.get("MSKU") or fields.get("SKU"): s += 2
            if fields.get("产品图") or fields.get("product_image"): s += 1
            if order_obj.order_status: s += 1
            if order_obj.tracking_no: s += 1
            return s

        rows_sorted = sorted(rows, key=lambda r: (score(r), r.updated_at or r.created_at or datetime.min, r.id), reverse=True)
        keeper = rows_sorted[0]
        dupes = rows_sorted[1:]

        keeper_ext = crud.get_order_ext(db, keeper.id)
        keeper_fields = dict(keeper_ext.fields or {}) if keeper_ext and isinstance(keeper_ext.fields, dict) else {}
        for d in dupes:
            dext = crud.get_order_ext(db, d.id)
            dfields = dext.fields if dext and isinstance(dext.fields, dict) else {}
            for k, v in (dfields or {}).items():
                if keeper_fields.get(k) in (None, "", " ", "None") and v not in (None, "", " ", "None"):
                    keeper_fields[k] = v
            if not keeper.tracking_no and d.tracking_no:
                keeper.tracking_no = d.tracking_no
            if not keeper.order_status and d.order_status:
                keeper.order_status = d.order_status
        if keeper_fields:
            crud.upsert_order_ext_bulk(db, keeper.id, keeper_fields)

        for d in dupes:
            db.query(models.SupplierQuoteResponse).filter(models.SupplierQuoteResponse.quote_request_id.in_(
                db.query(models.SupplierQuoteRequest.id).filter(models.SupplierQuoteRequest.internal_order_id == d.id)
            )).delete(synchronize_session=False)
            db.query(models.SupplierQuoteRequest).filter(models.SupplierQuoteRequest.internal_order_id == d.id).delete(synchronize_session=False)
            db.query(models.KapiExportItem).filter(models.KapiExportItem.internal_order_id == d.id).delete(synchronize_session=False)
            db.query(models.InternalOrderItem).filter(models.InternalOrderItem.internal_order_id == d.id).delete(synchronize_session=False)
            db.query(models.InternalOrderPackage).filter(models.InternalOrderPackage.internal_order_id == d.id).delete(synchronize_session=False)
            db.query(models.InternalOrderExt).filter(models.InternalOrderExt.internal_order_id == d.id).delete(synchronize_session=False)
            db.query(models.InternalOrder).filter(models.InternalOrder.id == d.id).delete(synchronize_session=False)
            removed_ids.append(d.id)
        kept += 1
    db.commit()
    return {"ok": True, "groups": groups_count, "kept": kept, "removed": len(removed_ids), "removed_ids": removed_ids[:50]}


@router.post("/sync-kapi-sign")
def sync_kapi_sign(payload: dict, db: Session = Depends(get_db)):
    order_ids = payload.get("order_ids") or []
    kcfg = crud.get_config(db, "kapi")
    kcfgv = kcfg.config_value if kcfg and isinstance(kcfg.config_value, dict) else {}
    api_key = str(payload.get("api_key") or kcfgv.get("api_key") or "").strip()
    base_url = str(payload.get("base_url") or kcfgv.get("base_url") or "").strip() or "https://tran.wedoexpress.com"
    if not order_ids:
        raise HTTPException(status_code=400, detail="missing order_ids")
    if not api_key:
        raise HTTPException(status_code=400, detail="missing api_key")

    pairs = []
    missing_kapi_no = []
    for oid in order_ids:
        order = crud.get_internal_order(db, int(oid))
        if not order:
            continue
        ext = crud.get_order_ext(db, order.id)
        fields = ext.fields if ext and ext.fields else {}
        order_no = _extract_kapi_order_no(fields)
        if not order_no:
            missing_kapi_no.append({"id": order.id, "platform_order_no": order.platform_order_no})
            continue
        pairs.append((order.id, order_no))

    uniq_order_nos = sorted(list({x[1] for x in pairs}))
    if not uniq_order_nos:
        return {
            "ok": False,
            "updated": 0,
            "missing_kapi_no": len(missing_kapi_no),
            "details": [],
            "missing_examples": missing_kapi_no[:10],
        }

    url = _build_kapi_url(base_url, "/openApi/truck/getOrders")
    headers = {"Api-Key": api_key, "Content-Type": "application/json"}
    try:
        resp = requests.post(url, headers=headers, json={"orderNos": uniq_order_nos}, timeout=30)
        try:
            data = resp.json()
        except Exception:
            body = (resp.text or "")[:500]
            raise HTTPException(status_code=400, detail=f"kapi getOrders non-json response, status={resp.status_code}, url={url}, body={body}")
        if not isinstance(data, dict):
            raise HTTPException(status_code=400, detail=f"kapi getOrders invalid response type: {type(data).__name__}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"kapi request failed: {e}")

    if not isinstance(data, dict) or not data.get("success"):
        return {
            "ok": False,
            "updated": 0,
            "missing_kapi_no": len(missing_kapi_no),
            "response": data,
            "missing_examples": missing_kapi_no[:10],
        }

    by_no = {str(x.get("orderNo") or "").strip(): x for x in (data.get("data") or [])}
    updated = 0
    details = []
    for oid, ono in pairs:
        row = by_no.get(ono)
        if not row:
            details.append({"id": oid, "orderNo": ono, "status": "not_found_in_response"})
            continue
        pod = row.get("podUrl") or ""
        track = row.get("trackNumber") or ""
        ext_update = {
            "卡派后台单号": f"{ono}#",
            "orderNo#": f"{ono}#",
            "卡派跟踪号": track,
            "卡派状态": row.get("status"),
            "卡派isReady": row.get("isReady"),
            "卡派PRO": row.get("proNumber") or "",
            "卡派LabelUrl": row.get("labelUrl") or "",
            "卡派BOLUrl": row.get("bolUrl") or "",
            "卡派错误": row.get("error") or "",
        }
        if track:
            ext_update["单号"] = ext_update.get("单号") or track
        crud.upsert_order_ext_bulk(db, oid, ext_update)
        updated += 1
        details.append(
            {
                "id": oid,
                "orderNo": ono,
                "isReady": row.get("isReady"),
                "status": row.get("status"),
                "trackNumber": track,
                "podUrl": pod,
            }
        )
    return {
        "ok": True,
        "updated": updated,
        "requested": len(uniq_order_nos),
        "missing_kapi_no": len(missing_kapi_no),
        "missing_examples": missing_kapi_no[:10],
        "details": details[:30],
    }


@router.post("/kapi-query-freight")
def kapi_query_freight(payload: dict, db: Session = Depends(get_db)):
    kcfg = crud.get_config(db, "kapi")
    kcfgv = kcfg.config_value if kcfg and isinstance(kcfg.config_value, dict) else {}
    api_key = str(payload.get("api_key") or kcfgv.get("api_key") or "").strip()
    base_url = str(payload.get("base_url") or kcfgv.get("base_url") or "").strip() or "https://tran.wedoexpress.com"
    request_data = _sanitize_kapi_query_payload(payload.get("request_data") or {})
    if not api_key:
        raise HTTPException(status_code=400, detail="missing api_key")
    if not isinstance(request_data, dict):
        raise HTTPException(status_code=400, detail="invalid request_data")

    headers = {"Api-Key": api_key, "Content-Type": "application/json"}
    is_express_payload = bool(request_data.get("carrierCode")) or isinstance((request_data.get("orderInfo") or {}).get("packages"), list)
    candidate_paths = ["/openApi/order/queryFreight", "/openApi/truck/queryFreight"] if is_express_payload else ["/openApi/truck/queryFreight", "/openApi/order/queryFreight"]
    out = None
    data = None
    used_path = ""
    attempts = []
    for p in candidate_paths:
        url = _build_kapi_url(base_url, p)
        try:
            resp = requests.post(url, headers=headers, json=request_data, timeout=40)
            raw = (resp.text or "")[:500]
            try:
                candidate = resp.json()
            except Exception:
                attempts.append({"path": p, "status": resp.status_code, "non_json": raw})
                continue
            if not isinstance(candidate, dict):
                attempts.append({"path": p, "status": resp.status_code, "invalid_type": type(candidate).__name__})
                continue
            out = candidate
            data = out.get("data")
            used_path = p
            break
        except Exception as e:
            attempts.append({"path": p, "error": str(e)})
            continue
    if out is None:
        raise HTTPException(status_code=400, detail={"message": "kapi queryFreight failed", "attempts": attempts})

    # truck: data is object with feeList; express: data is fee array
    fee_list = []
    if isinstance(data, dict):
        fee_list = data.get("feeList") or []
    elif isinstance(data, list):
        fee_list = data
    normalized = []
    for x in fee_list or []:
        if isinstance(data, list):
            normalized.append(
                {
                    "transportName": x.get("carrierCode"),
                    "carrierName": x.get("carrierCode"),
                    "carrierCode": x.get("carrierCode"),
                    "quoteNo": x.get("quoteNo"),
                    "orderNo": x.get("orderNo"),
                    "totalCharge": x.get("fee"),
                    "lineCharge": x.get("fee"),
                    "fuelCharge": None,
                    "accessorialCharge": None,
                    "insuranceCharge": None,
                    "transitDays": None,
                    "maxTransitDays": None,
                    "needReQuote": None,
                    "error": x.get("error"),
                    "currency": x.get("currency"),
                    "packageNum": x.get("packageNum"),
                }
            )
        else:
            normalized.append(
                {
                    "transportName": x.get("transportName"),
                    "carrierName": x.get("carrierName"),
                    "carrierCode": x.get("carrierCode"),
                    "quoteNo": x.get("quoteNo"),
                    "orderNo": x.get("orderNo"),
                    "totalCharge": x.get("totalCharge"),
                    "lineCharge": x.get("lineCharge"),
                    "fuelCharge": x.get("fuelCharge"),
                    "accessorialCharge": x.get("accessorialCharge"),
                    "insuranceCharge": x.get("insuranceCharge"),
                    "transitDays": x.get("transitDays"),
                    "maxTransitDays": x.get("maxTransitDays"),
                    "needReQuote": x.get("needReQuote"),
                    "error": x.get("error"),
                }
            )
    return {
        "ok": bool(isinstance(out, dict) and out.get("success")),
        "response": out,
        "source_path": used_path,
        "orderNo": (data or {}).get("orderNo") if isinstance(data, dict) else "",
        "isCompleted": (data or {}).get("isCompleted") if isinstance(data, dict) else None,
        "feeList": normalized,
    }


@router.post("/kapi-create-order")
def kapi_create_order(payload: dict, db: Session = Depends(get_db)):
    kcfg = crud.get_config(db, "kapi")
    kcfgv = kcfg.config_value if kcfg and isinstance(kcfg.config_value, dict) else {}
    api_key = str(payload.get("api_key") or kcfgv.get("api_key") or "").strip()
    base_url = str(payload.get("base_url") or kcfgv.get("base_url") or "").strip() or "https://tran.wedoexpress.com"
    request_data = payload.get("request_data") or {}
    order_id = payload.get("order_id")
    if not api_key:
        raise HTTPException(status_code=400, detail="missing api_key")
    if not isinstance(request_data, dict):
        raise HTTPException(status_code=400, detail="invalid request_data")
    if not request_data.get("orderNo") or not request_data.get("quoteNo"):
        raise HTTPException(status_code=400, detail="missing orderNo/quoteNo")

    url = _build_kapi_url(base_url, "/openApi/truck/createOrder")
    headers = {"Api-Key": api_key, "Content-Type": "application/json"}
    try:
        resp = requests.post(url, headers=headers, json=request_data, timeout=40)
        try:
            out = resp.json()
        except Exception:
            body = (resp.text or "")[:500]
            raise HTTPException(status_code=400, detail=f"kapi createOrder non-json response, status={resp.status_code}, url={url}, body={body}")
        if not isinstance(out, dict):
            raise HTTPException(status_code=400, detail=f"kapi createOrder invalid response type: {type(out).__name__}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"kapi createOrder failed: {e}")

    data = out.get("data") if isinstance(out, dict) else {}
    if bool(isinstance(out, dict) and out.get("success")) and isinstance(data, dict):
        track = data.get("trackNumber") or ""
        kapi_no = data.get("orderNo") or request_data.get("orderNo") or ""
        if order_id:
            ext_update = {
                "卡派后台单号": f"{kapi_no}#" if kapi_no else "",
                "orderNo#": f"{kapi_no}#" if kapi_no else "",
                "卡派跟踪号": track,
                "单号": track or "",
                "卡派isReady": data.get("isReady"),
                "卡派quoteNo": request_data.get("quoteNo"),
            }
            crud.upsert_order_ext_bulk(db, int(order_id), ext_update)

    return {
        "ok": bool(isinstance(out, dict) and out.get("success")),
        "response": out,
        "orderNo": (data or {}).get("orderNo") if isinstance(data, dict) else "",
        "trackNumber": (data or {}).get("trackNumber") if isinstance(data, dict) else "",
        "isReady": (data or {}).get("isReady") if isinstance(data, dict) else None,
    }


@router.post("/{internal_order_id}/quote-request")
def create_quote_request(internal_order_id: int, db: Session = Depends(get_db)):
    order = crud.get_internal_order(db, internal_order_id)
    if not order:
        raise HTTPException(status_code=404, detail="order not found")
    quote_no = f"Q{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    visible_payload = build_supplier_visible_payload(db, internal_order_id)
    req = crud.create_quote_request(db, internal_order_id, quote_no, visible_payload)
    return {"quote_no": req.quote_no, "status": req.quote_status}


@router.post("/{internal_order_id}/sync-lingxing")
def sync_order_to_lingxing(internal_order_id: int, db: Session = Depends(get_db)):
    order = crud.get_internal_order(db, internal_order_id)
    ext = crud.get_order_ext(db, internal_order_id)
    if not order or not ext:
        raise HTTPException(status_code=404, detail="order/ext not found")

    cfg = get_lingxing_config(db)
    app_id = cfg.get("app_id")
    app_secret = cfg.get("app_secret")
    if not app_id or not app_secret:
        raise HTTPException(status_code=400, detail="missing app_id/app_secret")

    token = get_access_token(app_id, app_secret)
    if token.get("code") not in (200, "200"):
        return token
    access_token = token.get("data", {}).get("access_token")

    fields = ext.fields or {}
    global_order_no = fields.get("global_order_no") or fields.get("lingxing_order_number")
    if not global_order_no:
        raise HTTPException(status_code=400, detail="missing global_order_no")

    address_info = {
        "address_line1": fields.get("address_line1") or "",
        "address_line2": fields.get("address_line2") or "",
        "city": fields.get("city") or "",
        "postal_code": fields.get("postal_code") or "",
        "receiver_country_code": fields.get("country_code") or "",
        "receiver_mobile": fields.get("receiver_mobile") or "",
        "receiver_name": fields.get("receiver_name") or "",
        "receiver_tel": fields.get("receiver_tel") or "",
        "state_or_region": fields.get("state_or_region") or "",
    }

    order_item_list = []
    if fields.get("sku") or fields.get("purchase_qty") or fields.get("unit_price"):
        price = fields.get("unit_price")
        if price is not None:
            try:
                price = int(float(price) * 1000000)
            except Exception:
                price = None
        order_item_list.append({
            "sku": fields.get("sku"),
            "quantity": fields.get("purchase_qty"),
            "price": price,
            "type": 3,
        })

    payload = [{
        "global_order_no": int(global_order_no),
        "address_info": address_info,
        "order_item_list": order_item_list or [],
    }]

    return update_fbm_order(access_token, app_id, payload)


@router.post("/fetch-images")
def fetch_images_by_asin(payload: dict, db: Session = Depends(get_db)):
    order_ids = payload.get("order_ids") or []
    if not order_ids:
        raise HTTPException(status_code=400, detail="missing order_ids")
    cfg = get_lingxing_config(db)
    app_id = cfg.get("app_id")
    app_secret = cfg.get("app_secret")
    if not app_id or not app_secret:
        raise HTTPException(status_code=400, detail="missing app_id/app_secret")
    token = get_access_token(app_id, app_secret)
    if token.get("code") not in (200, "200"):
        raise HTTPException(status_code=400, detail=str(token))
    access_token = token.get("data", {}).get("access_token")
    # build asin list from ext
    asin_to_orders = {}
    sku_to_orders = {}
    for oid in order_ids:
        ext = crud.get_order_ext(db, oid)
        asin = (ext.fields.get("asin") if ext and ext.fields else None)
        sku = (ext.fields.get("sku") if ext and ext.fields else None)
        if asin:
            asin_to_orders.setdefault(asin, []).append(oid)
        elif sku:
            sku_to_orders.setdefault(sku, []).append(oid)

    if not asin_to_orders and not sku_to_orders:
        return {"updated": 0}

    updated = 0
    errors = []
    if asin_to_orders:
        sid_value = resolve_sid_list(access_token, app_id, cfg.get("sid_list") or cfg.get("listing_sid"))
        payload_req = {
            "sid": sid_value,
            "is_pair": 1,
            "is_delete": 0,
            "search_field": "asin",
            "search_value": list(asin_to_orders.keys()),
            "exact_search": 1,
            "store_type": 1,
            "offset": 0,
            "length": 100,
        }
        res = get_listing_search(access_token, app_id, payload_req)
        if res.get("code") != 0:
            errors.append(res)
            res = {"data": []}
        for row in res.get("data", []) or []:
            asin = row.get("asin")
            img = row.get("small_image_url")
            if not asin or not img:
                continue
            for oid in asin_to_orders.get(asin, []):
                ext_update = {"product_image": img}
                if row.get("seller_sku"):
                    ext_update["sku"] = row.get("seller_sku")
                if row.get("local_name"):
                    ext_update["product_name"] = row.get("local_name")
                if row.get("asin"):
                    ext_update["asin"] = row.get("asin")
                crud.upsert_order_ext_bulk(db, oid, ext_update)
                items = crud.get_order_items(db, oid)
                if items:
                    items[0].product_image = img
                    db.commit()
                updated += 1
    if sku_to_orders:
        payload_req = {
            "sid": sid_value,
            "is_pair": 1,
            "is_delete": 0,
            "search_field": "seller_sku",
            "search_value": list(sku_to_orders.keys()),
            "exact_search": 1,
            "store_type": 1,
            "offset": 0,
            "length": 100,
        }
        res = get_listing_search(access_token, app_id, payload_req)
        if res.get("code") != 0:
            errors.append(res)
            res = {"data": []}
        for row in res.get("data", []) or []:
            sku = row.get("seller_sku") or row.get("local_sku")
            img = row.get("small_image_url")
            if not sku or not img:
                continue
            for oid in sku_to_orders.get(sku, []):
                ext_update = {"product_image": img}
                if row.get("asin"):
                    ext_update["asin"] = row.get("asin")
                if row.get("local_name"):
                    ext_update["product_name"] = row.get("local_name")
                if row.get("seller_sku"):
                    ext_update["sku"] = row.get("seller_sku")
                crud.upsert_order_ext_bulk(db, oid, ext_update)
                items = crud.get_order_items(db, oid)
                if items:
                    items[0].product_image = img
                    db.commit()
                updated += 1
    # stats
    with_asin = 0
    with_img = 0
    with_sku = 0
    for oid in order_ids:
        ext = crud.get_order_ext(db, oid)
        if ext and ext.fields:
            if ext.fields.get("asin"):
                with_asin += 1
            if ext.fields.get("sku"):
                with_sku += 1
            img = ext.fields.get("product_image")
            if img and img != "/":
                with_img += 1
    return {"updated": updated, "errors": errors[:3], "with_asin": with_asin, "with_sku": with_sku, "with_img": with_img}


@router.post("/backfill-detail")
def backfill_from_order_detail(payload: dict, db: Session = Depends(get_db)):
    order_ids = payload.get("order_ids") or []
    if not order_ids:
        # auto pick all orders
        order_ids = [o.id for o in db.query(models.InternalOrder).all()]
    cfg = get_lingxing_config(db)
    app_id = cfg.get("app_id")
    app_secret = cfg.get("app_secret")
    if not app_id or not app_secret:
        raise HTTPException(status_code=400, detail="missing app_id/app_secret")
    token = get_access_token(app_id, app_secret)
    if token.get("code") not in (200, "200"):
        raise HTTPException(status_code=400, detail=str(token))
    access_token = token.get("data", {}).get("access_token")

    updated = 0
    errors = []
    # batch by 200 order ids
    from app.integrations.lingxing_client import get_mws_order_detail
    platform_ids = []
    for oid in order_ids:
        order = crud.get_internal_order(db, oid)
        if order and order.platform_order_no:
            platform_ids.append(order.platform_order_no)
    for i in range(0, len(platform_ids), 200):
        batch = platform_ids[i:i+200]
        detail = get_mws_order_detail(access_token, app_id, batch)
        if detail.get("code") != 0:
            errors.append(detail)
            continue
        for d in detail.get("data", []) or []:
            amazon_order_id = d.get("amazon_order_id")
            order = crud.get_order_by_platform_no(db, amazon_order_id)
            if not order:
                continue
            item_list = d.get("item_list") or []
            if not item_list:
                continue
            it = item_list[0]
            ext_update = {}
            if it.get("asin"):
                ext_update["asin"] = it.get("asin")
            sku_val = it.get("sku") or it.get("seller_sku")
            if sku_val:
                ext_update["sku"] = sku_val
            if it.get("product_name") or it.get("title"):
                ext_update["product_name"] = it.get("product_name") or it.get("title")
            if d.get("latest_ship_date"):
                ext_update["latest_ship_date"] = d.get("latest_ship_date")
                ext_update["amz_ship"] = d.get("latest_ship_date")
            if d.get("earliest_delivery_date"):
                ext_update["earliest_delivery_date"] = d.get("earliest_delivery_date")
            if d.get("latest_delivery_date"):
                ext_update["latest_delivery_date"] = d.get("latest_delivery_date")
                ext_update["amz_deliver"] = f"{d.get('earliest_delivery_date','')} - {d.get('latest_delivery_date','')}".strip(" -")
            img = it.get("pic_url")
            if img and img != "/":
                ext_update["product_image"] = img
            if ext_update:
                crud.upsert_order_ext_bulk(db, order.id, ext_update)
            items = crud.get_order_items(db, order.id)
            if items:
                if items[0].sku in (None, "", " "):
                    items[0].sku = it.get("sku") or it.get("seller_sku")
                if items[0].product_name in (None, "", " "):
                    items[0].product_name = it.get("product_name") or it.get("title")
                if items[0].quantity in (None, 0):
                    items[0].quantity = it.get("quantity_ordered")
                if img and img != "/" and items[0].product_image in (None, "", "/"):
                    items[0].product_image = img
                db.commit()
            updated += 1

    # try listing search for missing images
    missing_asin = []
    missing_sku = []
    missing_sku_source = {}
    for oid in order_ids:
        ext = crud.get_order_ext(db, oid)
        if not ext or not ext.fields:
            continue
        if not ext.fields.get("product_image"):
            if ext.fields.get("asin"):
                missing_asin.append(ext.fields.get("asin"))
            elif ext.fields.get("sku"):
                val = ext.fields.get("sku")
                missing_sku.append(val)
                missing_sku_source.setdefault(val, []).append(oid)
    if missing_asin:
        sid_value = resolve_sid_list(access_token, app_id, cfg.get("sid_list") or cfg.get("listing_sid"))
        res = get_listing_search(access_token, app_id, {
            "sid": sid_value,
            "is_pair": 1,
            "is_delete": 0,
            "search_field": "asin",
            "search_value": missing_asin[:100],
            "exact_search": 1,
            "store_type": 1,
            "offset": 0,
            "length": 100,
        })
        if res.get("code") == 0:
            for row in res.get("data", []) or []:
                asin = row.get("asin")
                img = row.get("small_image_url")
                if not asin or not img:
                    continue
                for oid in order_ids:
                    ext = crud.get_order_ext(db, oid)
                    if ext and ext.fields and ext.fields.get("asin") == asin:
                        ext_update = {"product_image": img}
                        if row.get("seller_sku"):
                            ext_update["sku"] = row.get("seller_sku")
                        if row.get("local_name"):
                            ext_update["product_name"] = row.get("local_name")
                        if row.get("asin"):
                            ext_update["asin"] = row.get("asin")
                        crud.upsert_order_ext_bulk(db, oid, ext_update)
                        items = crud.get_order_items(db, oid)
                        if items:
                            items[0].product_image = img
                            db.commit()
                        updated += 1
        else:
            errors.append(res)
    if missing_sku:
        sid_value = resolve_sid_list(access_token, app_id, cfg.get("sid_list") or cfg.get("listing_sid"))
        res = get_listing_search(access_token, app_id, {
            "sid": sid_value,
            "is_pair": 1,
            "is_delete": 0,
            "search_field": "seller_sku",
            "search_value": missing_sku[:100],
            "exact_search": 1,
            "store_type": 1,
            "offset": 0,
            "length": 100,
        })
        if res.get("code") == 0:
            for row in res.get("data", []) or []:
                sku = row.get("seller_sku") or row.get("local_sku")
                img = row.get("small_image_url")
                if not sku or not img:
                    continue
                for oid in missing_sku_source.get(sku, []):
                    ext = crud.get_order_ext(db, oid)
                    if ext and ext.fields:
                        ext_update = {"product_image": img}
                        if row.get("seller_sku"):
                            ext_update["sku"] = row.get("seller_sku")
                        if row.get("local_name"):
                            ext_update["product_name"] = row.get("local_name")
                        if row.get("asin"):
                            ext_update["asin"] = row.get("asin")
                        crud.upsert_order_ext_bulk(db, oid, ext_update)
                        items = crud.get_order_items(db, oid)
                        if items:
                            items[0].product_image = img
                            db.commit()
                        updated += 1
        else:
            errors.append(res)
        # try msku/local_sku if seller_sku failed
        res2 = get_listing_search(access_token, app_id, {
            "sid": sid_value,
            "is_pair": 1,
            "is_delete": 0,
            "search_field": "msku",
            "search_value": missing_sku[:100],
            "exact_search": 1,
            "store_type": 1,
            "offset": 0,
            "length": 100,
        })
        if res2.get("code") == 0:
            for row in res2.get("data", []) or []:
                sku = row.get("seller_sku") or row.get("local_sku") or row.get("msku")
                img = row.get("small_image_url")
                if not sku or not img:
                    continue
                for oid in missing_sku_source.get(sku, []):
                    ext = crud.get_order_ext(db, oid)
                    if ext and ext.fields:
                        ext_update = {"product_image": img}
                        if row.get("seller_sku"):
                            ext_update["sku"] = row.get("seller_sku")
                        if row.get("local_name"):
                            ext_update["product_name"] = row.get("local_name")
                        if row.get("asin"):
                            ext_update["asin"] = row.get("asin")
                        crud.upsert_order_ext_bulk(db, oid, ext_update)
                        items = crud.get_order_items(db, oid)
                        if items:
                            items[0].product_image = img
                            db.commit()
                        updated += 1

    return {"updated": updated, "errors": errors[:3]}
